import sys
import re
import os
import openpyxl
from datetime import datetime, date
import calendar

# ─────────────────────────────────────────────────────────────────────
# PATHS  — update each month as needed
# ─────────────────────────────────────────────────────────────────────
COMPLIANCE_PATH     = r'C:\Automation\portal_Compliance_Model.xlsx'
SERVICE_TICKET_PATH = r'C:\Automation\2026_portal_Service_Ticket_Report.xlsx'
MAPS_FOLDER         = r'C:\Automation\Maps'
AUDIT_MONTH         = 'January 2026'   # ← update each month

# ─────────────────────────────────────────────────────────────────────
# AUDIT PERIOD SETUP
# ─────────────────────────────────────────────────────────────────────
def get_rules(audit_month_str):
    dt = datetime.strptime(audit_month_str, "%B %Y")
    month_end = date(dt.year, dt.month, calendar.monthrange(dt.year, dt.month)[1])
    if   dt.month in (1,2,3):  qn, qy = 4, dt.year - 1
    elif dt.month in (4,5,6):  qn, qy = 1, dt.year
    elif dt.month in (7,8,9):  qn, qy = 2, dt.year
    else:                      qn, qy = 3, dt.year
    q_months = {
        1: ('JAN','FEB','MAR','JANUARY','FEBRUARY','MARCH'),
        2: ('APR','MAY','JUN','APRIL','MAY','JUNE'),
        3: ('JUL','AUG','SEP','JULY','AUGUST','SEPTEMBER'),
        4: ('OCT','NOV','DEC','OCTOBER','NOVEMBER','DECEMBER'),
    }[qn]
    q_words = {
        1: ['Q1','Q-1','1ST QTR','1ST QUARTER','FIRST QUARTER'],
        2: ['Q2','Q-2','2ND QTR','2ND QUARTER','SECOND QUARTER'],
        3: ['Q3','Q-3','3RD QTR','3RD QUARTER','THIRD QUARTER'],
        4: ['Q4','Q-4','4TH QTR','4TH QUARTER','FOURTH QUARTER'],
    }[qn]
    return {
        'annual_year': dt.year,
        'month_end':   month_end,
        'qn': qn, 'qy': qy,
        'q_months': q_months,
        'q_words':  q_words,
    }

RR = get_rules(AUDIT_MONTH)
print(f"Audit month   : {AUDIT_MONTH}")
print(f"Valid through : {RR['month_end']}")
print(f"Quarter needed: Q{RR['qn']} {RR['qy']}")

# ─────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────
def norm_addr(addr):
    if not addr: return ''
    addr = str(addr).upper().strip()
    addr = re.sub(r'[.,]', '', addr)
    return re.sub(r'\s+', ' ', addr).strip()

def norm_name(name):
    if not name: return ''
    name = str(name).upper().strip()
    name = re.sub(r'[.,\-]', ' ', name)
    return re.sub(r'\s+', ' ', name).strip()

def parse_date(v):
    if v is None: return None
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date):     return v
    s = str(v).strip()
    if not s or s.upper() in ('N/A','NONE','NAN',''): return None
    for fmt in ('%m/%d/%Y','%m/%d/%y','%Y-%m-%d'):
        try: return datetime.strptime(s, fmt).date()
        except: pass
    return None

def exp_status(raw):
    d = parse_date(raw)
    if d is None or d.year <= 1900: return 'MANUAL_REVIEW_DATE'
    if d < RR['month_end']:         return 'EXPIRED'
    return 'VALID'

def has_year(title, yr):
    return str(yr) in str(title or '').upper()

def has_target_quarter(title):
    t = str(title or '').upper()
    qy = str(RR['qy'])
    has_q = any(w in t for w in RR['q_words'])
    years = re.findall(r'\b(20\d{2}|\d{2})\b', t)
    norm  = {('20'+y if len(y)==2 else y) for y in years}
    if has_q and (not norm or qy in norm): return True
    mm = sum(1 for m in RR['q_months'] if m in t)
    if mm >= 2 and qy in t: return True
    return False

def has_explicit_wrong_year(title):
    """True if title has a quarter keyword AND an explicit year that is NOT the target."""
    t = str(title or '').upper()
    has_q = any(w in t for w in RR['q_words'])
    if not has_q: return False
    years = re.findall(r'\b(20\d{2})\b', t)
    norm = set(years)
    qy = str(RR['qy'])
    return bool(norm) and qy not in norm

def name_tokens(name):
    parts = [p for p in norm_name(name).split() if len(p) > 1]
    if len(parts) >= 2: return parts[0], parts[-1]
    if len(parts) == 1: return parts[0], None
    return None, None

def row_tech_matches(row_tech, svc_tech):
    if not row_tech or not svc_tech: return False
    rf, rl = name_tokens(row_tech)
    sf, sl = name_tokens(svc_tech)
    if not rf or not sf: return False
    if rl and sl: return rf == sf and rl == sl
    return rf == sf

def tech_in_title(tech, title):
    if not tech or not title: return False
    tn = norm_name(tech); ti = norm_name(title)
    parts = tn.split()
    if len(parts) >= 2: return parts[0] in ti and parts[-1] in ti
    return tn in ti

def is_vague(t):
    return str(t or '').strip().upper() in {
        'CERTIFICATE','CERTIFICATION','LICENSE','LICENCE','LICENSES',
        'client_a','client_aCO','IPM','GMP','CGMP','PURDUE','AIB','NPMA',
        'INSURANCE','TRAINING','N/A','GOLD MEDAL','NPMA'}

# ─────────────────────────────────────────────────────────────────────
# MAP EVALUATION
# ─────────────────────────────────────────────────────────────────────
# Equipment keyword patterns — all matched case-insensitively on OCR text
# ERB = Exterior Rodent Bait Station
ERB_PATTERNS = [
    r'(?:ERB|EXTERIOR\s+RODENT(?:\s+BAIT)?(?:\s+STATIONS?)?'
    r'|BAIT\s+STATIONS?|EXT\.?\s*RODENT|EXTERIOR\s+BAIT)(?:\s+COUNT)?[:\s]+(\d+)',
]
# IRT = Interior Rodent Trap
IRT_PATTERNS = [
    r'(?:IRT|INTERIOR\s+RODENT(?:\s+TRAPS?)?'
    r'|TIN\s+CATS?|INT\.?\s*RODENT|INTERIOR\s+TRAP'
    r'|MOUSE\s+SNAP\s+TRAP|RAT\s+SNAP\s+TRAP)(?:\s+COUNT)?[:\s]+(\d+)',
]
# IFL = Insect/Fly Light Trap
IFL_PATTERNS = [
    r'(?:IFL|ILT|FLY\s+LIGHTS?|INSECT\s+LIGHT(?:\s+TRAPS?)?'
    r'|LIGHT\s+TRAPS?)(?:\s+COUNT)?[:\s]+(\d+)',
]
# Date patterns
DATE_PATTERNS = [
    r'\b\d{1,2}/\d{1,2}/\d{4}\b',                        # MM/DD/YYYY
    r'\b\d{1,2}/\d{1,2}/\d{2}\b',                        # MM/DD/YY
    r'\b(?:JAN(?:UARY)?|FEB(?:RUARY)?|MAR(?:CH)?|APR(?:IL)?'
    r'|MAY|JUN(?:E)?|JUL(?:Y)?|AUG(?:UST)?|SEP(?:TEMBER)?'
    r'|OCT(?:OBER)?|NOV(?:EMBER)?|DEC(?:EMBER)?)'
    r'\.?\s+\d{1,2},?\s+\d{4}\b',                        # January 2, 2024
]
# Signature/reviewer patterns
SIGNED_PATTERNS = [
    r'MAPREVIEW',
    r'\bBY[:\s]+[A-Z]',
    r'REVIEWED\s+BY',
    r'REVISED\s+BY',
    r'TECH(?:NICIAN)?[:\s]+[A-Z]',
    r'ACCOUNT\s+MANAGER',
    r'SERVICE\s+MANAGER',
    r'SIGNATURE',
    r'\b[A-Z][a-z]+\s+[A-Z][a-z]+\b',   # any "Firstname Lastname" pattern
]

def _ocr_pdf_first_page(pdf_path):
    """OCR first page of a PDF. Returns uppercased text or '' on failure."""
    try:
        import pytesseract
        from PIL import Image
        import tempfile, subprocess as sp
        with tempfile.TemporaryDirectory() as tmpdir:
            prefix = os.path.join(tmpdir, 'p')
            sp.run(['pdftoppm', '-jpeg', '-r', '200', '-f', '1', '-l', '1',
                    pdf_path, prefix], capture_output=True, timeout=30)
            imgs = sorted(f for f in os.listdir(tmpdir) if f.endswith('.jpg'))
            if not imgs:
                return ''
            text = pytesseract.image_to_string(
                Image.open(os.path.join(tmpdir, imgs[0])), config='--psm 6')
            return text.upper()
    except Exception:
        return ''

def _find_count(patterns, text):
    """Return first integer match from patterns in text, or None."""
    for pat in patterns:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            try:
                return int(m.group(1))
            except Exception:
                pass
    return None

def evaluate_map(address, sow_erb=None, sow_irt=None, sow_ifl=None):
    """
    Look for a map PDF in MAPS_FOLDER, OCR it, extract counts,
    check signed/dated, compare to SOW.
    Returns dict of all map_* fields.
    """
    blank = {
        'map_found': 'NO', 'map_file': '', 'map_page_count': '',
        'map_erb': '', 'map_irt': '', 'map_ifl': '',
        'map_signed': 'MANUAL_REVIEW', 'map_dated': 'MANUAL_REVIEW',
        'map_status': 'MISSING', 'map_flags': [],
    }

    # ── find PDF ──────────────────────────────────────────────────────
    # Try exact match first, then prefix/partial matching as fallback
    pdf_path = os.path.join(MAPS_FOLDER, f'{norm_addr(address)}_map.pdf')
    if not os.path.exists(pdf_path):
        target = norm_addr(address).lower()
        pdf_path = None
        try:
            for fname in os.listdir(MAPS_FOLDER):
                if fname.lower().endswith('_map.pdf'):
                    stem = fname.lower().replace('_map.pdf', '')
                    # Match exact, OR address starts with stem (file has short address)
                    # OR stem starts with address (file has longer address)
                    if (stem == target
                            or target.startswith(stem)
                            or stem.startswith(target)):
                        pdf_path = os.path.join(MAPS_FOLDER, fname)
                        break
        except Exception:
            pass
        if not pdf_path:
            blank['map_flags'].append(f'Map PDF not found — looked for: {norm_addr(address)[:40]}')
            return blank

    r = dict(blank)
    r['map_found']  = 'YES'
    r['map_file']   = os.path.basename(pdf_path)
    r['map_status'] = 'MANUAL_REVIEW'   # default until proven otherwise
    r['map_flags']  = []

    # ── page count ───────────────────────────────────────────────────
    try:
        from pypdf import PdfReader
        r['map_page_count'] = len(PdfReader(pdf_path).pages)
    except Exception:
        r['map_page_count'] = 'UNKNOWN'

    # ── OCR ──────────────────────────────────────────────────────────
    text = _ocr_pdf_first_page(pdf_path)
    if not text.strip():
        r['map_flags'].append('PDF text unreadable — manual review required')
        return r   # all map fields stay MANUAL_REVIEW/blank

    # ── parse counts ─────────────────────────────────────────────────
    erb = _find_count(ERB_PATTERNS, text)
    irt = _find_count(IRT_PATTERNS, text)
    ifl = _find_count(IFL_PATTERNS, text)

    r['map_erb'] = erb if erb is not None else ''
    r['map_irt'] = irt if irt is not None else ''
    r['map_ifl'] = ifl if ifl is not None else ''

    if erb is None: r['map_flags'].append('ERB count not readable')
    if irt is None: r['map_flags'].append('IRT count not readable')
    if ifl is None: r['map_flags'].append('IFL count not readable')

    # ── dated ────────────────────────────────────────────────────────
    dated = any(re.search(p, text, re.IGNORECASE) for p in DATE_PATTERNS)
    r['map_dated'] = 'PASS' if dated else 'MANUAL_REVIEW'
    if not dated:
        r['map_flags'].append('No date found on map')

    # ── signed ───────────────────────────────────────────────────────
    signed = any(re.search(p, text) for p in SIGNED_PATTERNS)
    r['map_signed'] = 'PASS' if signed else 'MANUAL_REVIEW'
    if not signed:
        r['map_flags'].append('No signature/reviewer found on map')

    # ── compare counts to SOW ────────────────────────────────────────
    count_mismatches = []
    if sow_erb is not None and erb is not None and erb != int(sow_erb):
        count_mismatches.append(f'ERB: Map={erb} SOW={sow_erb}')
    if sow_irt is not None and irt is not None and irt != int(sow_irt):
        count_mismatches.append(f'IRT: Map={irt} SOW={sow_irt}')
    if sow_ifl is not None and ifl is not None and ifl != int(sow_ifl):
        count_mismatches.append(f'IFL: Map={ifl} SOW={sow_ifl}')
    r['map_flags'].extend(count_mismatches)

    # ── overall map_status ───────────────────────────────────────────
    any_count_found  = (erb is not None or irt is not None or ifl is not None)
    counts_match     = any_count_found and len(count_mismatches) == 0
    everything_clean = (counts_match and r['map_signed'] == 'PASS'
                        and r['map_dated'] == 'PASS'
                        and erb is not None and irt is not None and ifl is not None)
    r['map_status'] = 'PASS' if everything_clean else 'MANUAL_REVIEW'

    return r

def is_bundled(t):
    t = str(t or '').upper()
    return any(kw in t for kw in ['ALL APPLICATOR','ALL TECH','ALL LICENSE',
        'ALL CERTS','MULTIPLE','CERTIFICATES AND LICENSES'])

# ─────────────────────────────────────────────────────────────────────
# DOC CLASSIFIER
# ─────────────────────────────────────────────────────────────────────
# Common typo corrections — applied before any keyword matching
TYPO_MAP = {
    # client_a
    'client_a':'client_a', 'client_a':'client_a', 'client_a':'client_a',
    # License
    'LISENSE':'LICENSE', 'LICESNSE':'LICENSE', 'LISCENSE':'LICENSE',
    'LICSENSE':'LICENSE', 'LICNESE':'LICENSE', 'LISENSCE':'LICENSE',
    'LISENSE':'LICENSE', 'LICNSE':'LICENSE',
    # Certificate
    'CERTFICATE':'CERTIFICATE', 'CERTIFCATE':'CERTIFICATE',
    'CERIFICATE':'CERTIFICATE', 'CERFICATE':'CERTIFICATE',
    'CERTIFICTE':'CERTIFICATE', 'CERTIFIATE':'CERTIFICATE',
    # Insurance
    'INSURNACE':'INSURANCE', 'INSURACE':'INSURANCE',
    'INSUANCE':'INSURANCE', 'INSURENCE':'INSURANCE',
    # Pesticide
    'PESTICLDE':'PESTICIDE', 'PESTICDE':'PESTICIDE',
    'PESTICIED':'PESTICIDE', 'PESTIICIDE':'PESTICIDE',
    # Assessment
    'ASSESMENT':'ASSESSMENT', 'ASSEMENT':'ASSESSMENT',
    'ASESSMENT':'ASSESSMENT', 'ASSESEMENT':'ASSESSMENT',
    # Annual
    'ANNAUL':'ANNUAL', 'ANUAL':'ANNUAL', 'ANNUALL':'ANNUAL',
    # Quarterly
    'QUATERLY':'QUARTERLY', 'QUARERLY':'QUARTERLY',
    'QUARTERY':'QUARTERLY', 'QTERLY':'QUARTERLY',
    # Facility
    'FACILTY':'FACILITY', 'FACILTIY':'FACILITY', 'FACILITY':'FACILITY',
    # Certification
    'CERTIFCATION':'CERTIFICATION', 'CERTIFIATION':'CERTIFICATION',
    'CERTIFCIATION':'CERTIFICATION',
    # Applicator
    'APLICATOR':'APPLICATOR', 'APPLCATOR':'APPLICATOR',
    'APLICATER':'APPLICATOR',
    # Business
    'BUSINES':'BUSINESS', 'BUSSINESS':'BUSINESS',
}

def fix_typos(t):
    """Replace known misspellings in a title string before keyword matching."""
    for wrong, right in TYPO_MAP.items():
        t = t.replace(wrong, right)
    return t

def classify(title, doc_area):
    """Returns doc type string or None.
    doc_area is used to help disambiguate (insurance vs branch_techs vs qa_manual)."""
    if not title or str(title).strip().upper() in ('N/A','NONE',''): return None
    t = fix_typos(str(title).upper().strip())
    area = str(doc_area or '').strip().lower()

    # BRANCH LICENSE — insurance section only for ambiguous titles
    BRANCH_KW = [
        'BRANCH LIC','BRANCH LICENSE','PESTICIDE BUSINESS LICENSE',
        'PEST CONTROL BUSINESS LICENSE','BUSINESS LICENSE','COMPANY LICENSE',
        'BRANCH PC LICENSE','company BRANCH LICENSE','FDACS BRANCH',
        'STRUCTURAL PEST CONTROL COMPANY LICENSE','SPCS BUSINESS LICENSE',
        'BUS LIC','BRANCH DEPT OF AG','PESTICIDE CONTRACTOR',
        'OCCUPATIONAL LICENSE','BUSINESS TAX RECEIPT','BROWARD',
        'AGENTS LICENSE','AGENTS CARDS','AGENT CARD',
    ]
    for kw in BRANCH_KW:
        if kw in t: return 'PC_License_Branch'
    # "License" or "Renewal Receipt" in insurance section = branch license
    if area == 'insurance' and ('LICENSE' in t or 'RENEWAL RECEIPT' in t):
        return 'PC_License_Branch'

    # COI
    COI_KW = ['COI','CERTIFICATE OF INSURANCE','CERTIFICATE OF LIABILITY',
        'LIABILITY INSURANCE','INSURANCE CERTIFICATE','INSURANCE 2026',
        'INSURANCE 2025','company COI','AUTO INSURANCE',
        '2026 INSURANCE','2025 INSURANCE']
    for kw in COI_KW:
        if kw in t: return 'COI'
    # "Insurance YYYY" in insurance section
    if area == 'insurance' and 'INSURANCE' in t:
        return 'COI'

    # ANNUAL REPORT
    ANNUAL_KW = ['ANNUAL ASSESSMENT','ANNUAL FACILITY ASSESSMENT',
        'ANNUAL SITE ASSESSMENT','ANNUAL INSPECTION','ANNUAL RISK ASSESSMENT',
        'ANNUAL FACILITY RISK','ANNUAL PLAN','ANNUAL QA','YEARLY ASSESSMENT',
        'FACILITY SITE ASSESSMENT','COMPLIANCE ASSESSMENT','RISK ASSESSMENT',
        'FLOOR LEVEL INSPECTION','QA INSPECTION REPORT','QA ASSESSMENT',
        'client_a SERVICE COMPLIANCE','client_b ANNUAL','MANAGER ANNUAL',
        'company ANNUAL','company FACILITY','SITE INSPECTION','SITE ASSESSMENT',
        'INSPECTION REPORT','ASSESMENT','ASSESSMENT']
    for kw in ANNUAL_KW:
        if kw in t: return 'Annual_Report'

    # QUARTERLY TREND
    TREND_KW = ['TREND','TRENDING','QUARTERLY TREND','PEST TREND',
        'PEST AUDIT TREND','PEST AUDIT REPORT','QUARTERLY ASSESSMENT',
        'QUARTERLY INSPECTION','QUARTERLY SERVICE','QUARTERLY QA',
        'QTRLY PEST AUDIT','PEST CAPTURES',
        'Q1 TREND','Q2 TREND','Q3 TREND','Q4 TREND',
        'Q-1 TREND','Q-2 TREND','Q-3 TREND','Q-4 TREND',
        '1ST QTR TREND','2ND QTR TREND','3RD QTR TREND','4TH QTR TREND',
        '1ST QUARTER PEST','2ND QUARTER PEST','3RD QUARTER PEST',
        '4TH QUARTER PEST','YEAR TREND','YEARLY TREND']
    for kw in TREND_KW:
        if kw in t: return 'Quarterly_Trend'

    # PESTICIDE LOG
    LOG_KW = ['PESTICIDE USAGE LOG','PESTICIDE USAGE REPORT',
        'PESTICIDE USAGE','PEST USAGE LOG','PEST USAGE',
        'PUL ','client_a BAINBRIDGE PUL','client_b BIN PESTICIDE',
        'PREVIOUS 12 MONTH PESTICIDE','PESTICIDE CONTRACTORS LICENSE',
        'CHEMICAL CORRECTION',
        'Q1 PESTICIDE USAGE','Q2 PESTICIDE USAGE',
        'Q3 PESTICIDE USAGE','Q4 PESTICIDE USAGE',
        '1ST QUARTER PESTICIDE','2ND QUARTER PESTICIDE',
        '3RD QUARTER PESTICIDE','4TH QUARTER PESTICIDE',
        'Q4 PESTICLDE','PESTICIDE LOG','PESTICIDE LOG']
    for kw in LOG_KW:
        if kw in t: return 'Pesticide_Log'

    # IPM CERT (tech-level)
    IPM_KW = ['IPM','GMP','CGMP','IMP','PPGM','PURDUE','GOLD MEDAL',
        'NPMA','AIB','PRECISION PROTECTION','FOOD PROCESSING',
        'FOOD SAFETY','SQF TRAINING','TEXAS A&M','CLEMSON','ACE CERTIFICATE',
        'ADVANCE IPM','ADVANCED IPM','ANNUAL IPM']
    for kw in IPM_KW:
        if kw in t: return 'IPM_Cert'

    # client_a CERT (tech-level)
    client_a_KW = ['client_a CERT','client_a CERT','client_a CERTIFICATE',
        'client_a CERTIFICATION','client_a CERTIFICATION','client_a TRAINING',
        'client_a TRAINING','client_a ','client_a SERVICE',
        'SERVICE AUDITS SALES','SERVICE, AUDITS, SALES',
        'SERVICING GOLD MEDAL','SERVICING PPGM','GOLD MEDAL TRAINING',
        'client_b CERT','client_b CGMP','client_a SERVICE',
        ]
    for kw in client_a_KW:
        if kw in t: return 'client_a_Cert'
    # "client_a" alone in branch_techs section = client_a cert (common portal naming)
    if area == 'branch_techs' and t.strip() in ('client_a','client_a'):
        return 'client_a_Cert'

    # TECH LICENSE
    TECH_KW = ['APPLICATOR LICENSE','APPLICATORS LICENSE','APPLICATOR CARD',
        'CERTIFIED APPLICATOR','TECHNICIAN LICENSE','SERVICE TECHNICIAN LICENSE',
        'TECH LICENSE','TECH CARD','STATE LICENSE','STATE LICENCE',
        'OPERATORS LICENSE','OPERATOR LICENSE','CERTIFIED OPERATOR',
        'CERTIFIED PEST CONTROL OPERATOR','PC LICENSE','PCO',
        'PESTICIDE LICENSE','PESTICIDE APPLICATOR LICENSE',
        'PEST CONTROL LICENSE','PEST LICENSE',
        'DEPT OF AG','DEPARTMENT OF AGRICULTURE',
        'TN LICENSE','LA LICENSE','WV APPLICATOR','MDA LICENSE',
        'ARKANSAS','AR AGENTS','OK LICENSE','RHODE ISLAND LICENSE',
        'FL LIC','FL LICENSE','SP LICENSE','SP/CO LICENSE',
        'IDPH','RECERTIFICATION','REGISTRATION',
        'MANAGER LICENSE','STRUCTURAL PEST CONTROL LICENSE',
        'STATE APPLICATOR','EMPLOYEE TRAING FILE','TRAINING FILE',
        'LICENSE CARD','LICENSE HOLDER','LICENSED UNDER',
        'ID CARD','STATE ID','MISS TECH ID CARD','AGRICULTURE CARD',
        '7C','ALL APPLICATOR CARDS','RECENT ADDITION APPLICATOR',
        'PRO LICENSE','GREENPRO','GREEN PRO','QUALITY PRO',
        'company LICENSE','APL','CERT CARD','CEUS',
        'COMMERCIAL OP','COMMERCIAL SERVICE CERTIFICATION']
    for kw in TECH_KW:
        if kw in t: return 'PC_License_Tech'
    # "License YYYY" in branch_techs = tech license
    if area == 'branch_techs' and 'LICENSE' in t:
        return 'PC_License_Tech'

    return None

# ─────────────────────────────────────────────────────────────────────
# LOAD WORKBOOK
# ─────────────────────────────────────────────────────────────────────
print(f"\nLoading: {COMPLIANCE_PATH}")
try:
    wb = openpyxl.load_workbook(COMPLIANCE_PATH)
    print("  ✓ Compliance model loaded")
except Exception as e:
    print(f"  ✗ FAILED to load compliance model: {e}")
    sys.exit(1)

# raw docs
ws_raw = wb['Raw_Document_Pull']
raw_docs = []
for i, row in enumerate(ws_raw.iter_rows(values_only=True)):
    if i == 0: continue
    if not row[0]: continue
    raw_docs.append({
        'address':     norm_addr(row[0]),
        'address_raw': row[0],
        'doc_area':    str(row[1] or '').strip().lower(),
        'tech_name':   str(row[2] or '').strip(),
        'doc_title':   str(row[3] or '').strip(),
        'exp_date':    row[4],
    })
print(f"  ✓ Raw docs: {len(raw_docs)} rows, {len(set(d['address'] for d in raw_docs))} unique addresses")

# last WO tech
ws_wo = wb['Last_WO_Tech']
last_wo_tech = {}
for i, row in enumerate(ws_wo.iter_rows(values_only=True)):
    if i == 0: continue
    if not row[0]: continue
    addr  = norm_addr(row[0])
    tech  = str(row[4] or '').strip()
    if tech.startswith('='): tech = ''
    last_wo_tech[addr] = tech
print(f"  ✓ Last WO Tech: {len(last_wo_tech)} entries")

# locations
locations = {}
for sheet in ['client_a_Locations','client_b_Locations']:
    ws = wb[sheet]
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0: continue
        if not row[0]: continue
        addr = norm_addr(row[0])
        locations[addr] = {
            'branch':       row[3],
            'state':        str(row[4] or '').upper().strip(),
            'company':      str(row[2] or ''),
            'is_franchise': str(row[5] or 'NO').upper() == 'YES',
        }
print(f"  ✓ Locations: {len(locations)} entries")

# state exceptions
ws_st   = wb['State_Exceptions']
state_exc = {}
for i, row in enumerate(ws_st.iter_rows(values_only=True)):
    if i == 0: continue
    if not row[0]: continue
    state_exc[str(row[0]).upper().strip()] = str(row[1] or 'YES').upper().strip()
print(f"  ✓ State exceptions: {len(state_exc)} states")

# service ticket (optional — won't crash if missing)
sow_counts     = {}
scanned_counts = {}
print(f"\nLoading service ticket report...")
try:
    wb_st = openpyxl.load_workbook(SERVICE_TICKET_PATH)
    ws_loc = wb_st['Location_Month_Summary']
    for i, row in enumerate(ws_loc.iter_rows(values_only=True)):
        if i == 0: continue
        if not row[0]: continue
        addr = norm_addr(row[0])
        sow_counts[addr] = {'ERB': row[5] or 0, 'IRT': row[6] or 0, 'IFL': row[7] or 0}
    ws_wos = wb_st['WO_Summary']
    for i, row in enumerate(ws_wos.iter_rows(values_only=True)):
        if i == 0: continue
        if not row[1]: continue
        addr = norm_addr(row[1])
        if addr not in scanned_counts:
            scanned_counts[addr] = {'ERB': 0, 'IRT': 0, 'IFL': 0}
        scanned_counts[addr]['ERB'] += (row[5] or 0)
        scanned_counts[addr]['IRT'] += (row[6] or 0)
        scanned_counts[addr]['IFL'] += (row[7] or 0)
    print(f"  ✓ Service ticket loaded — {len(sow_counts)} SOW rows")
except Exception as e:
    print(f"  ⚠ Service ticket not loaded: {e}")
    print("    Equipment match will show NO_DATA (this is OK, rest will still run)")

# group docs by address
docs_by_addr = {}
for d in raw_docs:
    docs_by_addr.setdefault(d['address'], []).append(d)

# ─────────────────────────────────────────────────────────────────────
# EVALUATE EACH LOCATION
# ─────────────────────────────────────────────────────────────────────
print(f"\nEvaluating {len(docs_by_addr)} locations...")
results = []

for addr, docs in docs_by_addr.items():
    loc        = locations.get(addr)
    tech_name  = last_wo_tech.get(addr, '')
    state      = (loc or {}).get('state', '')
    branch     = (loc or {}).get('branch', '')
    company    = (loc or {}).get('company', '')
    is_fra     = (loc or {}).get('is_franchise', False)
    if branch and str(branch).startswith('9'): is_fra = True
    state_req  = state_exc.get(state, 'YES') == 'YES'

    if not loc:
        print(f"  WARN: {addr} not in client_a or client_b locations tab")

    # starting values
    pc_branch  = 'MISSING'
    pc_tech    = 'MISSING' if state_req else 'EXEMPT'
    client_a_cert = 'MISSING'
    ipm_cert   = 'MISSING'
    coi        = 'MISSING'
    annual     = 'MISSING'
    trend      = 'EXEMPT' if is_fra else 'MISSING'
    pest_log   = 'EXEMPT' if is_fra else 'MISSING'
    doc_flags  = []

    for doc in docs:
        title    = doc['doc_title']
        doc_area = doc['doc_area']
        row_tech = doc['tech_name']
        es       = exp_status(doc['exp_date'])
        dtype    = classify(title, doc_area)

        if dtype is None:
            continue

        # ── BRANCH LICENSE ───────────────────────────────────────────
        if dtype == 'PC_License_Branch':
            if   es == 'VALID'              and pc_branch != 'PASS':   pc_branch = 'PASS'
            elif es == 'MANUAL_REVIEW_DATE' and pc_branch != 'PASS':
                pc_branch = 'MANUAL_REVIEW'
                doc_flags.append(f'PC_License_Branch: date unclear "{title}" exp={doc["exp_date"]}')
            elif es == 'EXPIRED'            and pc_branch != 'PASS':
                pc_branch = 'EXPIRED'
                doc_flags.append(f'PC_License_Branch: EXPIRED "{title}"')

        # ── COI ──────────────────────────────────────────────────────
        # NOTE: portal expiration dates are unreliable (~15% wrong).
        # Rule: if the title clearly contains the audit year (e.g. "Insurance 2026"),
        # that overrides the expiration date and counts as PASS.
        # If no year in title, we still check the expiration date as a fallback.
        elif dtype == 'COI':
            if has_year(title, RR['annual_year']):
                # Title has 2026 → trust the title, not the date
                coi = 'PASS'
            elif has_year(title, RR['annual_year'] - 1):
                # Title clearly says last year (e.g. "2025 COI") → wrong year
                if coi not in ('PASS',):
                    coi = 'WRONG_YEAR'
                doc_flags.append(f'COI: title shows {RR["annual_year"]-1}, need {RR["annual_year"]} "{title}"')
            else:
                # No year in title — fall back to expiration date check
                if es == 'VALID':
                    exp_d = parse_date(doc['exp_date'])
                    if exp_d and exp_d.year == RR['annual_year']:
                        coi = 'PASS'   # exp year matches audit year — trust it
                    elif coi not in ('PASS',):
                        coi = 'MANUAL_REVIEW'
                        doc_flags.append(f'COI: no year in title, exp date used - manual verify "{title}"')
                elif es == 'EXPIRED' and coi != 'PASS':
                    coi = 'EXPIRED'
                    doc_flags.append(f'COI: no year in title and EXPIRED "{title}" exp={doc["exp_date"]}')
                elif es == 'MANUAL_REVIEW_DATE' and coi != 'PASS':
                    coi = 'MANUAL_REVIEW'
                    doc_flags.append(f'COI: no year in title, date unclear "{title}" exp={doc["exp_date"]}')

        # ── ANNUAL REPORT ────────────────────────────────────────────
        elif dtype == 'Annual_Report':
            # only count qa_manual rows for annual/trend/log
            if doc_area != 'qa_manual': continue
            if has_year(title, RR['annual_year']):
                annual = 'PASS'
            elif annual != 'PASS':
                annual = 'MANUAL_REVIEW'
                doc_flags.append(f'Annual_Report: year unclear, need {RR["annual_year"]} "{title}"')

        # ── QUARTERLY TREND ──────────────────────────────────────────
        elif dtype == 'Quarterly_Trend':
            if doc_area != 'qa_manual' or trend == 'EXEMPT': continue
            if has_target_quarter(title):
                trend = 'PASS'
            elif has_explicit_wrong_year(title):
                doc_flags.append(f'Quarterly_Trend: wrong year, need Q{RR["qn"]} {RR["qy"]} "{title}"')
            elif trend != 'PASS':
                trend = 'MANUAL_REVIEW'
                doc_flags.append(f'Quarterly_Trend: need Q{RR["qn"]} {RR["qy"]} "{title}"')

        # ── PESTICIDE LOG ────────────────────────────────────────────
        elif dtype == 'Pesticide_Log':
            if doc_area != 'qa_manual' or pest_log == 'EXEMPT': continue
            if has_target_quarter(title):
                pest_log = 'PASS'
            elif has_explicit_wrong_year(title):
                doc_flags.append(f'Pesticide_Log: wrong year, need Q{RR["qn"]} {RR["qy"]} "{title}"')
            elif pest_log != 'PASS':
                pest_log = 'MANUAL_REVIEW'
                doc_flags.append(f'Pesticide_Log: need Q{RR["qn"]} {RR["qy"]} "{title}"')

        # ── TECH-LEVEL DOCS ──────────────────────────────────────────
        elif dtype in ('PC_License_Tech', 'client_a_Cert', 'IPM_Cert'):
            if doc_area == 'qa_manual': continue
            # Agents/all-tech cards in insurance section count as branch-level tech license
            if dtype == 'PC_License_Tech' and doc_area == 'insurance':
                if es == 'VALID' and pc_tech != 'PASS':
                    pc_tech = 'PASS'
                continue
            if not tech_name: continue
            name_match = row_tech_matches(row_tech, tech_name) or tech_in_title(tech_name, title)

            if dtype == 'PC_License_Tech' and state_req:
                if name_match:
                    if   es == 'VALID':              pc_tech = 'PASS'
                    elif es == 'MANUAL_REVIEW_DATE' and pc_tech != 'PASS':
                        pc_tech = 'MANUAL_REVIEW'
                        doc_flags.append(f'PC_License_Tech: date unclear for {tech_name} "{title}" exp={doc["exp_date"]}')
                    elif es == 'EXPIRED' and pc_tech != 'PASS':
                        pc_tech = 'EXPIRED'
                        doc_flags.append(f'PC_License_Tech: EXPIRED for {tech_name} "{title}"')
                elif is_bundled(title):
                    doc_flags.append(f'PC_License_Tech: bundled doc, manual review "{title}"')
                elif not is_vague(title):
                    doc_flags.append(f'PC_License_Tech: no match for {tech_name} "{title}"')

            elif dtype == 'client_a_Cert':
                if name_match:
                    client_a_cert = 'PASS'
                elif not is_vague(title):
                    doc_flags.append(f'client_a_Cert: no match for {tech_name} "{title}"')

            elif dtype == 'IPM_Cert':
                if name_match:
                    if has_year(title, RR['annual_year']):
                        ipm_cert = 'PASS'
                    elif has_year(title, RR['annual_year'] - 1):
                        if ipm_cert != 'PASS': ipm_cert = 'WRONG_YEAR'
                        doc_flags.append(f'IPM_Cert: 2025 found, need 2026 "{title}"')
                    else:
                        if ipm_cert != 'PASS': ipm_cert = 'MANUAL_REVIEW'
                        doc_flags.append(f'IPM_Cert: year unclear "{title}"')
                elif not is_vague(title):
                    doc_flags.append(f'IPM_Cert: no match for {tech_name} "{title}"')

    # ── EQUIPMENT MATCH ──────────────────────────────────────────────
    sow     = sow_counts.get(addr)
    scanned = scanned_counts.get(addr)
    if sow and scanned:
        erb_ok = sow['ERB'] == scanned['ERB']
        irt_ok = sow['IRT'] == scanned['IRT']
        ifl_ok = sow['IFL'] == scanned['IFL']
        equip  = 'MATCH' if (erb_ok and irt_ok and ifl_ok) else 'MISMATCH'
        if not erb_ok: doc_flags.append(f'ERB: SOW={sow["ERB"]} Scanned={scanned["ERB"]}')
        if not irt_ok: doc_flags.append(f'IRT: SOW={sow["IRT"]} Scanned={scanned["IRT"]}')
        if not ifl_ok: doc_flags.append(f'IFL: SOW={sow["IFL"]} Scanned={scanned["IFL"]}')
    else:
        equip = 'NO_DATA'

    # ── COMPLIANCE % ─────────────────────────────────────────────────
    checks = {
        'PC_License_Branch': pc_branch, 'PC_License_Tech': pc_tech,
        'client_a_Cert': client_a_cert,        'IPM_Cert': ipm_cert,
        'COI': coi,                      'Annual_Report': annual,
        'Quarterly_Trend': trend,        'Pesticide_Log': pest_log,
    }
    applicable     = [k for k,v in checks.items() if v != 'EXEMPT']
    passed         = [k for k in applicable if checks[k] == 'PASS']
    checks_app     = len(applicable)
    checks_pass    = len(passed)
    pct            = round(checks_pass / checks_app * 100, 1) if checks_app else 0

    # build flags string
    flag_issues = [f'{k}: {v}' for k,v in checks.items() if v in ('MISSING','EXPIRED','WRONG_YEAR','MANUAL_REVIEW')]
    if equip == 'MISMATCH': flag_issues.append('EQUIPMENT COUNT MISMATCH')
    flag_issues.extend(doc_flags)
    flags_str = ' | '.join(flag_issues) if flag_issues else 'NONE'

    original_addr = docs[0]['address_raw'] if docs else addr

    results.append({
        'audit_month':       AUDIT_MONTH,
        'address':           original_addr,
        'branch':            branch,
        'state':             state,
        'company':           company,
        'tech_name':         tech_name,
        'PC_license_branch': pc_branch,
        'PC_license_tech':   pc_tech,
        'client_a_cert':        client_a_cert,
        'IPM_cert':          ipm_cert,
        'COI':               coi,
        'annual_report':     annual,
        'quarterly_trend':   trend,
        'pesticide_log':     pest_log,
        'equipment_match':   equip,
        # ── map evaluation ──────────────────────────────────────────
        **{k: (v if not isinstance(v, list) else ' | '.join(v) if v else '')
           for k, v in evaluate_map(
               addr,
               sow_erb=sow_counts.get(addr, {}).get('ERB'),
               sow_irt=sow_counts.get(addr, {}).get('IRT'),
               sow_ifl=sow_counts.get(addr, {}).get('IFL'),
           ).items()},
        'checks_applicable': checks_app,
        'checks_passed':     checks_pass,
        'compliance_pct':    f'{pct}%',
        'flags':             flags_str,
        'last_checked':      datetime.today().strftime('%m/%d/%Y'),
    })
    print(f"  ✓ {original_addr}  {pct}%  tech=[{tech_name}]  branch_lic={pc_branch}  tech_lic={pc_tech}  client_a={client_a_cert}  ipm={ipm_cert}  coi={coi}  annual={annual}  trend={trend}  log={pest_log}")

# ─────────────────────────────────────────────────────────────────────
# WRITE TO COMPLIANCE_EVAL
# ─────────────────────────────────────────────────────────────────────
HEADERS = [
    'audit_month','address','branch','state','company','tech_name',
    'PC_license_branch','PC_license_tech','client_a_cert','IPM_cert',
    'COI','annual_report','quarterly_trend','pesticide_log',
    'equipment_match','map_found','map_file','map_page_count',
    'map_erb','map_irt','map_ifl','map_signed','map_dated','map_status',
    'map_flags','checks_applicable','checks_passed','compliance_pct',
    'flags','last_checked'
]

print(f"\nWriting {len(results)} rows to Compliance_Eval...")
ws_eval = wb['Compliance_Eval']

# clear old data (keep row 1 for headers)
for row in ws_eval.iter_rows(min_row=2):
    for cell in row:
        cell.value = None

# write headers row 1
for col, h in enumerate(HEADERS, 1):
    ws_eval.cell(row=1, column=col, value=h)

# write data rows
for row_idx, result in enumerate(results, 2):
    for col_idx, key in enumerate(HEADERS, 1):
        ws_eval.cell(row=row_idx, column=col_idx, value=result.get(key, ''))

wb.save(COMPLIANCE_PATH)
print(f"✓ Saved. {len(results)} locations written to Compliance_Eval.")
print("\nDone!")
