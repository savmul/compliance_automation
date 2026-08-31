import sys
import re
import os
import openpyxl
from datetime import datetime, date
import calendar

# ─────────────────────────────────────────────────────────────────────
# NOTE ON GENERICIZATION
# Client, employer, and system names have been replaced throughout this
# file for public release: CLIENT_A / CLIENT_B are two national accounts,
# and "the portal" is a proprietary service-management platform.
#
# The document-title keyword lists below are therefore illustrative
# rather than runnable — in production they contain the real strings that
# appear in scanned document titles. The matching logic, thresholds, and
# control flow are unchanged.
# ─────────────────────────────────────────────────────────────────────

# ─────────────────────────────────────────────────────────────────────
# PATHS  — update each month as needed
# ─────────────────────────────────────────────────────────────────────
COMPLIANCE_PATH     = r'C:\Automation\Compliance_Model.xlsx'
SERVICE_TICKET_PATH = r'C:\Automation\Service_Ticket_Report.xlsx'
MAPS_FOLDER         = r'C:\Automation\Maps'
AUDIT_MONTH         = 'August 2026'   # ← update each month

# ─────────────────────────────────────────────────────────────────────
# AUDIT PERIOD SETUP
# ─────────────────────────────────────────────────────────────────────
def get_rules(audit_month_str):
    dt = datetime.strptime(audit_month_str, "%B %Y")
    month_end = date(dt.year, dt.month, calendar.monthrange(dt.year, dt.month)[1])
    if   dt.month in (1,2,3):  qn, qy = 4, dt.year - 1
    elif dt.month in (4,5,6):  qn, qy = 1, dt.year
    elif dt.month in (7,8,9):  qn, qy = 2, dt.year
    else:                      qn, qy = 3, dt.year
    q_months = {
        1: ('JAN','FEB','MAR','JANUARY','FEBRUARY','MARCH'),
        2: ('APR','MAY','JUN','APRIL','MAY','JUNE'),
        3: ('JUL','AUG','SEP','JULY','AUGUST','SEPTEMBER'),
        4: ('OCT','NOV','DEC','OCTOBER','NOVEMBER','DECEMBER'),
    }[qn]
    q_words = {
        1: ['Q1','Q-1','1ST QTR','1ST QUARTER','FIRST QUARTER'],
        2: ['Q2','Q-2','2ND QTR','2ND QUARTER','SECOND QUARTER'],
        3: ['Q3','Q-3','3RD QTR','3RD QUARTER','THIRD QUARTER'],
        4: ['Q4','Q-4','4TH QTR','4TH QUARTER','FOURTH QUARTER'],
    }[qn]
    return {
        'annual_year': dt.year,
        'month_end':   month_end,
        'qn': qn, 'qy': qy,
        'q_months': q_months,
        'q_words':  q_words,
    }

RR = get_rules(AUDIT_MONTH)
print(f"Audit month   : {AUDIT_MONTH}")
print(f"Valid through : {RR['month_end']}")
print(f"Quarter needed: Q{RR['qn']} {RR['qy']}")

# ─────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────
_SUFFIX = {'ROAD':'RD','STREET':'ST','AVENUE':'AVE','DRIVE':'DR','BOULEVARD':'BLVD',
           'LANE':'LN','HIGHWAY':'HWY','CIRCLE':'CIR','COURT':'CT','PARKWAY':'PKWY',
           'PLACE':'PL','TERRACE':'TER','ROUTE':'RT'}

def norm_addr(addr):
    """Normalize an address into a join key. Same logic as compare_equipment.py
    so SOW_Lookup / WO_Counts_Detail / Map_Manual_Data / Raw_Document_Pull /
    Client_A_Locations all collapse to the same key.
    - keeps only the street portion before the first comma
    - drops any prefix text before the first street-number token
    - standardizes ROAD/STREET/AVENUE/... -> RD/ST/AVE/...
    """
    if not addr: return ''
    s = str(addr).split(',')[0].upper().replace('.', '')
    s = re.sub(r'\s+', ' ', s).strip()
    toks = s.split()
    for i, t in enumerate(toks):
        if t[:1].isdigit():
            toks = toks[i:]
            break
    return ' '.join(_SUFFIX.get(w, w) for w in toks)

def norm_name(name):
    if not name: return ''
    name = str(name).upper().strip()
    name = re.sub(r'[.,\-]', ' ', name)
    return re.sub(r'\s+', ' ', name).strip()

def parse_date(v):
    if v is None: return None
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date):     return v
    s = str(v).strip()
    if not s or s.upper() in ('N/A','NONE','NAN',''): return None
    for fmt in ('%m/%d/%Y','%m/%d/%y','%Y-%m-%d'):
        try: return datetime.strptime(s, fmt).date()
        except: pass
    return None

def exp_status(raw):
    d = parse_date(raw)
    if d is None or d.year <= 1900: return 'MANUAL_REVIEW_DATE'
    if d < RR['month_end']:         return 'EXPIRED'
    return 'VALID'

def has_year(title, yr):
    return str(yr) in str(title or '').upper()

def has_target_quarter(title):
    t = str(title or '').upper()
    qy = str(RR['qy'])
    has_q = any(w in t for w in RR['q_words'])
    years = re.findall(r'\b(20\d{2}|\d{2})\b', t)
    norm  = {('20'+y if len(y)==2 else y) for y in years}
    if has_q and (not norm or qy in norm): return True
    mm = sum(1 for m in RR['q_months'] if m in t)
    if mm >= 2 and qy in t: return True
    return False

def has_explicit_wrong_year(title):
    """True if title has a quarter keyword AND an explicit year that is NOT the target."""
    t = str(title or '').upper()
    has_q = any(w in t for w in RR['q_words'])
    if not has_q: return False
    years = re.findall(r'\b(20\d{2})\b', t)
    norm = set(years)
    qy = str(RR['qy'])
    return bool(norm) and qy not in norm

def parse_inspection_date(s, audit_year):
    """Parse inspection date string into a date object.
    Handles: '01/03/2025', '27 Jan', '6 Feb (4 weeks ago)', '15 Jan 2026'
    Falls back to audit_year when no year present."""
    if not s: return None
    s = str(s).strip()
    # Remove relative parts like "(4 weeks ago)"
    s = re.sub(r'\(.*?\)', '', s).strip()
    # Try standard formats first
    for fmt in ('%m/%d/%Y', '%m/%d/%y', '%Y-%m-%d', '%d %b %Y', '%d %B %Y'):
        try: return datetime.strptime(s, fmt).date()
        except: pass
    # Try partial "27 Jan" or "6 Feb" — assume audit_year
    for fmt in ('%d %b', '%d %B'):
        try:
            d = datetime.strptime(s, fmt)
            return d.replace(year=audit_year).date()
        except: pass
    return None

def implied_quarter_year(title, insp_date):
    """For a doc with quarter keyword but no year, use inspection_date to infer year.
    Rule: Q4 docs uploaded in Jan-Mar → Q4 of previous year.
    All other quarter/month combos → same year as inspection_date."""
    if not insp_date: return None
    t = str(title or '').upper()
    # Find quarter number from title
    qn = None
    if any(w in t for w in ['Q4','Q-4','4TH QTR','4TH QUARTER','FOURTH QUARTER']): qn = 4
    elif any(w in t for w in ['Q1','Q-1','1ST QTR','1ST QUARTER','FIRST QUARTER']): qn = 1
    elif any(w in t for w in ['Q2','Q-2','2ND QTR','2ND QUARTER','SECOND QUARTER']): qn = 2
    elif any(w in t for w in ['Q3','Q-3','3RD QTR','3RD QUARTER','THIRD QUARTER']): qn = 3
    if qn is None: return None
    # Q4 uploaded in Jan-Mar means it covers Q4 of prior year
    if qn == 4 and insp_date.month in (1, 2, 3):
        return (qn, insp_date.year - 1)
    return (qn, insp_date.year)

def name_tokens(name):
    parts = [p for p in norm_name(name).split() if len(p) > 1]
    if len(parts) >= 2: return parts[0], parts[-1]
    if len(parts) == 1: return parts[0], None
    return None, None

def row_tech_matches(row_tech, svc_tech):
    if not row_tech or not svc_tech: return False
    rf, rl = name_tokens(row_tech)
    sf, sl = name_tokens(svc_tech)
    if not rf or not sf: return False
    if rl and sl: return rf == sf and rl == sl
    return rf == sf

def tech_in_title(tech, title):
    if not tech or not title: return False
    tn = norm_name(tech); ti = norm_name(title)
    parts = tn.split()
    if len(parts) >= 2: return parts[0] in ti and parts[-1] in ti
    return tn in ti

def is_vague(t):
    return str(t or '').strip().upper() in {
        'CERTIFICATE','CERTIFICATION','LICENSE','LICENCE','LICENSES',
        'CLIENT_A','CLIENT_A','IPM','GMP','CGMP','PURDUE','AIB','NPMA',
        'INSURANCE','TRAINING','N/A','GOLD MEDAL','NPMA'}

# ─────────────────────────────────────────────────────────────────────
# MAP EVALUATION
# ─────────────────────────────────────────────────────────────────────
# Equipment keyword patterns — all matched case-insensitively on OCR text
# ERB = Exterior Rodent Bait Station
ERB_PATTERNS = [
    r'(?:ERB|EXTERIOR\s+RODENT(?:\s+BAIT)?(?:\s+STATIONS?)?'
    r'|BAIT\s+STATIONS?|EXT\.?\s*RODENT|EXTERIOR\s+BAIT)(?:\s+COUNT)?[:\s]+(\d+)',
]
# IRT = Interior Rodent Trap
IRT_PATTERNS = [
    r'(?:IRT|INTERIOR\s+RODENT(?:\s+TRAPS?)?'
    r'|TIN\s+CATS?|INT\.?\s*RODENT|INTERIOR\s+TRAP'
    r'|MOUSE\s+SNAP\s+TRAP|RAT\s+SNAP\s+TRAP)(?:\s+COUNT)?[:\s]+(\d+)',
]
# IFL = Insect/Fly Light Trap
IFL_PATTERNS = [
    r'(?:IFL|ILT|FLY\s+LIGHTS?|INSECT\s+LIGHT(?:\s+TRAPS?)?'
    r'|LIGHT\s+TRAPS?)(?:\s+COUNT)?[:\s]+(\d+)',
]
# Date patterns
DATE_PATTERNS = [
    r'\b\d{1,2}/\d{1,2}/\d{4}\b',                        # MM/DD/YYYY
    r'\b\d{1,2}/\d{1,2}/\d{2}\b',                        # MM/DD/YY
    r'\b(?:JAN(?:UARY)?|FEB(?:RUARY)?|MAR(?:CH)?|APR(?:IL)?'
    r'|MAY|JUN(?:E)?|JUL(?:Y)?|AUG(?:UST)?|SEP(?:TEMBER)?'
    r'|OCT(?:OBER)?|NOV(?:EMBER)?|DEC(?:EMBER)?)'
    r'\.?\s+\d{1,2},?\s+\d{4}\b',                        # January 2, 2024
]
# Signature/reviewer patterns
SIGNED_PATTERNS = [
    r'MAPREVIEW',
    r'\bBY[:\s]+[A-Z]',
    r'REVIEWED\s+BY',
    r'REVISED\s+BY',
    r'TECH(?:NICIAN)?[:\s]+[A-Z]',
    r'ACCOUNT\s+MANAGER',
    r'SERVICE\s+MANAGER',
    r'SIGNATURE',
    r'\b[A-Z][a-z]+\s+[A-Z][a-z]+\b',   # any "Firstname Lastname" pattern
]

def _ocr_pdf_first_page(pdf_path):
    """OCR first page of a PDF. Returns uppercased text or '' on failure."""
    try:
        import pytesseract
        from PIL import Image
        import tempfile, subprocess as sp
        with tempfile.TemporaryDirectory() as tmpdir:
            prefix = os.path.join(tmpdir, 'p')
            sp.run(['pdftoppm', '-jpeg', '-r', '200', '-f', '1', '-l', '1',
                    pdf_path, prefix], capture_output=True, timeout=30)
            imgs = sorted(f for f in os.listdir(tmpdir) if f.endswith('.jpg'))
            if not imgs:
                return ''
            text = pytesseract.image_to_string(
                Image.open(os.path.join(tmpdir, imgs[0])), config='--psm 6')
            return text.upper()
    except Exception:
        return ''

def _find_count(patterns, text):
    """Return first integer match from patterns in text, or None."""
    for pat in patterns:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            try:
                return int(m.group(1))
            except Exception:
                pass
    return None

def evaluate_map(address, sow_erb=None, sow_irt=None, sow_ifl=None):
    """
    Look for a map PDF in MAPS_FOLDER, OCR it, extract counts,
    check signed/dated, compare to SOW.
    Returns dict of all map_* fields.
    """
    blank = {
        'map_found': 'NO', 'map_file': '', 'map_page_count': '',
        'map_erb': '', 'map_irt': '', 'map_ifl': '',
        'map_signed': 'MANUAL_REVIEW', 'map_dated': 'MANUAL_REVIEW',
        'map_status': 'MISSING', 'map_flags': [],
    }

    # ── find PDF ──────────────────────────────────────────────────────
    # Try exact match first, then prefix/partial matching as fallback
    pdf_path = os.path.join(MAPS_FOLDER, f'{norm_addr(address)}_map.pdf')
    if not os.path.exists(pdf_path):
        target = norm_addr(address).lower()
        pdf_path = None
        try:
            for fname in os.listdir(MAPS_FOLDER):
                if fname.lower().endswith('_map.pdf'):
                    stem = fname.lower().replace('_map.pdf', '')
                    # Match exact, OR address starts with stem (file has short address)
                    # OR stem starts with address (file has longer address)
                    if (stem == target
                            or target.startswith(stem)
                            or stem.startswith(target)):
                        pdf_path = os.path.join(MAPS_FOLDER, fname)
                        break
        except Exception:
            pass
        if not pdf_path:
            blank['map_flags'].append(f'Map PDF not found — looked for: {norm_addr(address)[:40]}')
            return blank

    r = dict(blank)
    r['map_found']  = 'YES'
    r['map_file']   = os.path.basename(pdf_path)
    r['map_status'] = 'MANUAL_REVIEW'   # default until proven otherwise
    r['map_flags']  = []

    # ── page count ───────────────────────────────────────────────────
    try:
        from pypdf import PdfReader
        r['map_page_count'] = len(PdfReader(pdf_path).pages)
    except Exception:
        r['map_page_count'] = 'UNKNOWN'

    # ── OCR ──────────────────────────────────────────────────────────
    text = _ocr_pdf_first_page(pdf_path)
    if not text.strip():
        r['map_flags'].append('PDF text unreadable — manual review required')
        return r   # all map fields stay MANUAL_REVIEW/blank

    # ── parse counts ─────────────────────────────────────────────────
    erb = _find_count(ERB_PATTERNS, text)
    irt = _find_count(IRT_PATTERNS, text)
    ifl = _find_count(IFL_PATTERNS, text)

    r['map_erb'] = erb if erb is not None else ''
    r['map_irt'] = irt if irt is not None else ''
    r['map_ifl'] = ifl if ifl is not None else ''

    if erb is None: r['map_flags'].append('ERB count not readable')
    if irt is None: r['map_flags'].append('IRT count not readable')
    if ifl is None: r['map_flags'].append('IFL count not readable')

    # ── dated ────────────────────────────────────────────────────────
    dated = any(re.search(p, text, re.IGNORECASE) for p in DATE_PATTERNS)
    if dated:
        r['map_dated'] = 'PASS'
    elif text:  # OCR ran but found nothing
        r['map_dated'] = 'NO_OCR'
        r['map_flags'].append('Map date not readable — verify manually')
    else:  # no OCR available
        r['map_dated'] = 'NO_OCR'

    # ── signed ───────────────────────────────────────────────────────
    signed = any(re.search(p, text) for p in SIGNED_PATTERNS)
    if signed:
        r['map_signed'] = 'PASS'
    else:
        r['map_signed'] = 'NO_OCR'
        r['map_flags'].append('Map signature not readable — verify manually')

    # ── compare counts to SOW ────────────────────────────────────────
    count_mismatches = []
    if sow_erb is not None and erb is not None and erb != int(sow_erb):
        count_mismatches.append(f'ERB: Map={erb} SOW={sow_erb}')
    if sow_irt is not None and irt is not None and irt != int(sow_irt):
        count_mismatches.append(f'IRT: Map={irt} SOW={sow_irt}')
    if sow_ifl is not None and ifl is not None and ifl != int(sow_ifl):
        count_mismatches.append(f'IFL: Map={ifl} SOW={sow_ifl}')
    r['map_flags'].extend(count_mismatches)

    # ── overall map_status ───────────────────────────────────────────
    any_count_found  = (erb is not None or irt is not None or ifl is not None)
    counts_match     = any_count_found and len(count_mismatches) == 0
    everything_clean = (counts_match and r['map_signed'] == 'PASS'
                        and r['map_dated'] == 'PASS'
                        and erb is not None and irt is not None and ifl is not None)
    if everything_clean:
        r['map_status'] = 'PASS'
    elif r['map_dated'] == 'NO_OCR' or r['map_signed'] == 'NO_OCR':
        r['map_status'] = 'NO_OCR'
    else:
        r['map_status'] = 'MANUAL_REVIEW'

    return r

def is_bundled(t):
    t = str(t or '').upper()
    return any(kw in t for kw in ['ALL APPLICATOR','ALL TECH','ALL LICENSE',
        'ALL CERTS','MULTIPLE','CERTIFICATES AND LICENSES'])

def is_branch_tech_bundle(t):
    """True for docs that bundle a business/branch license AND a tech license
    in one PDF (e.g. 'Business & Tech License'). We can't confirm either side
    without opening it, so both columns get flagged for manual review."""
    t = str(t or '').upper()
    has_lic  = 'LICENSE' in t or 'LIC ' in t
    has_bus  = 'BUSINESS' in t or 'BUS ' in t or 'BRANCH' in t
    has_tech = 'TECH' in t or 'TECHNICIAN' in t
    return has_lic and has_bus and has_tech

# ─────────────────────────────────────────────────────────────────────
# DOC CLASSIFIER
# ─────────────────────────────────────────────────────────────────────
# Common typo corrections — applied before any keyword matching
TYPO_MAP = {
    # Client name misspellings (genericized; production list holds the real variants)
    'CLEINT':'CLIENT_A', 'CLINET':'CLIENT_A', 'CLIETN':'CLIENT_A',
    # License
    'LISENSE':'LICENSE', 'LICESNSE':'LICENSE', 'LISCENSE':'LICENSE',
    'LICSENSE':'LICENSE', 'LICNESE':'LICENSE', 'LISENSCE':'LICENSE',
    'LISENSE':'LICENSE', 'LICNSE':'LICENSE',
    # Certificate
    'CERTFICATE':'CERTIFICATE', 'CERTIFCATE':'CERTIFICATE',
    'CERIFICATE':'CERTIFICATE', 'CERFICATE':'CERTIFICATE',
    'CERTIFICTE':'CERTIFICATE', 'CERTIFIATE':'CERTIFICATE',
    # Insurance
    'INSURNACE':'INSURANCE', 'INSURACE':'INSURANCE',
    'INSUANCE':'INSURANCE', 'INSURENCE':'INSURANCE',
    # Pesticide
    'PESTICLDE':'PESTICIDE', 'PESTICDE':'PESTICIDE',
    'PESTICIED':'PESTICIDE', 'PESTIICIDE':'PESTICIDE',
    # Assessment
    'ASSESMENT':'ASSESSMENT', 'ASSEMENT':'ASSESSMENT',
    'ASESSMENT':'ASSESSMENT', 'ASSESEMENT':'ASSESSMENT',
    # Annual
    'ANNAUL':'ANNUAL', 'ANUAL':'ANNUAL', 'ANNUALL':'ANNUAL',
    # Quarterly
    'QUATERLY':'QUARTERLY', 'QUARERLY':'QUARTERLY',
    'QUARTERY':'QUARTERLY', 'QTERLY':'QUARTERLY',
    # Facility
    'FACILTY':'FACILITY', 'FACILTIY':'FACILITY', 'FACILITY':'FACILITY',
    # Certification
    'CERTIFCATION':'CERTIFICATION', 'CERTIFIATION':'CERTIFICATION',
    'CERTIFCIATION':'CERTIFICATION',
    # Applicator
    'APLICATOR':'APPLICATOR', 'APPLCATOR':'APPLICATOR',
    'APLICATER':'APPLICATOR',
    # Business
    'BUSINES':'BUSINESS', 'BUSSINESS':'BUSINESS',
}

def fix_typos(t):
    """Replace known misspellings in a title string before keyword matching."""
    for wrong, right in TYPO_MAP.items():
        t = t.replace(wrong, right)
    return t

def classify(title, doc_area):
    """Returns doc type string or None.
    doc_area is used to help disambiguate (insurance vs branch_techs vs qa_manual)."""
    if not title or str(title).strip().upper() in ('N/A','NONE',''): return None
    t = fix_typos(str(title).upper().strip())
    area = str(doc_area or '').strip().lower()

    # BRANCH LICENSE — insurance section only for ambiguous titles
    BRANCH_KW = [
        'BRANCH LIC','BRANCH LICENSE','PESTICIDE BUSINESS LICENSE',
        'PEST CONTROL BUSINESS LICENSE','BUSINESS LICENSE','COMPANY LICENSE',
        'BRANCH PC LICENSE','PROVIDER BRANCH LICENSE','FDACS BRANCH',
        'STRUCTURAL PEST CONTROL COMPANY LICENSE','SPCS BUSINESS LICENSE',
        'BUS LIC','BRANCH DEPT OF AG','PESTICIDE CONTRACTOR',
        'OCCUPATIONAL LICENSE','BUSINESS TAX RECEIPT','BROWARD',
        'AGENTS LICENSE','AGENTS CARDS','AGENT CARD',
    ]
    for kw in BRANCH_KW:
        if kw in t: return 'PC_License_Branch'
    # "License" or "Renewal Receipt" in insurance section = branch license
    if area == 'insurance' and ('LICENSE' in t or 'RENEWAL RECEIPT' in t):
        return 'PC_License_Branch'

    # COI
    COI_KW = ['COI','CERTIFICATE OF INSURANCE','CERTIFICATE OF LIABILITY',
        'LIABILITY INSURANCE','INSURANCE CERTIFICATE','INSURANCE 2026',
        'INSURANCE 2025','PROVIDER COI','AUTO INSURANCE',
        '2026 INSURANCE','2025 INSURANCE']
    for kw in COI_KW:
        if kw in t: return 'COI'
    # "Insurance YYYY" in insurance section
    if area == 'insurance' and 'INSURANCE' in t:
        return 'COI'

    # ANNUAL REPORT
    ANNUAL_KW = ['ANNUAL ASSESSMENT','ANNUAL FACILITY ASSESSMENT',
        'ANNUAL SITE ASSESSMENT','ANNUAL INSPECTION','ANNUAL RISK ASSESSMENT',
        'ANNUAL FACILITY RISK','ANNUAL PLAN','ANNUAL QA','YEARLY ASSESSMENT',
        'FACILITY SITE ASSESSMENT','COMPLIANCE ASSESSMENT','RISK ASSESSMENT',
        'FLOOR LEVEL INSPECTION','QA INSPECTION REPORT','QA ASSESSMENT',
        'CLIENT_A SERVICE COMPLIANCE','CLIENT_B ANNUAL','MANAGER ANNUAL',
        'PROVIDER ANNUAL','PROVIDER FACILITY','SITE INSPECTION','SITE ASSESSMENT',
        'INSPECTION REPORT','ASSESMENT','ASSESSMENT']
    for kw in ANNUAL_KW:
        if kw in t: return 'Annual_Report'

    # QUARTERLY TREND
    TREND_KW = ['TREND','TRENDING','QUARTERLY TREND','PEST TREND',
        'PEST AUDIT TREND','PEST AUDIT REPORT','QUARTERLY ASSESSMENT',
        'QUARTERLY INSPECTION','QUARTERLY SERVICE','QUARTERLY QA',
        'QTRLY PEST AUDIT','PEST CAPTURES',
        'Q1 TREND','Q2 TREND','Q3 TREND','Q4 TREND',
        'Q-1 TREND','Q-2 TREND','Q-3 TREND','Q-4 TREND',
        '1ST QTR TREND','2ND QTR TREND','3RD QTR TREND','4TH QTR TREND',
        '1ST QUARTER PEST','2ND QUARTER PEST','3RD QUARTER PEST',
        '4TH QUARTER PEST','YEAR TREND','YEARLY TREND']
    for kw in TREND_KW:
        if kw in t: return 'Quarterly_Trend'

    # PESTICIDE LOG
    LOG_KW = ['PESTICIDE USAGE LOG','PESTICIDE USAGE REPORT',
        'PESTICIDE USAGE','PEST USAGE LOG','PEST USAGE',
        'PUL ','CLIENT_A SITE PUL','CLIENT_B BIN PESTICIDE',
        'PREVIOUS 12 MONTH PESTICIDE','PESTICIDE CONTRACTORS LICENSE',
        'CHEMICAL CORRECTION',
        'Q1 PESTICIDE USAGE','Q2 PESTICIDE USAGE',
        'Q3 PESTICIDE USAGE','Q4 PESTICIDE USAGE',
        '1ST QUARTER PESTICIDE','2ND QUARTER PESTICIDE',
        '3RD QUARTER PESTICIDE','4TH QUARTER PESTICIDE',
        'Q4 PESTICLDE','PESTICIDE LOG','PESTICIDE LOG']
    for kw in LOG_KW:
        if kw in t: return 'Pesticide_Log'

    # IPM CERT (tech-level)
    IPM_KW = ['IPM','GMP','CGMP','IMP','PPGM','PURDUE','GOLD MEDAL',
        'NPMA','AIB','PRECISION PROTECTION','FOOD PROCESSING',
        'FOOD SAFETY','SQF TRAINING','TEXAS A&M','CLEMSON','ACE CERTIFICATE',
        'ADVANCE IPM','ADVANCED IPM','ANNUAL IPM']
    for kw in IPM_KW:
        if kw in t: return 'IPM_Cert'

    # CLIENT_A CERT (tech-level)
    CLIENT_CERT_KW = ['CLIENT_A CERT','CLIENT_A CERT','CLIENT_A CERTIFICATE',
        'CLIENT_A CERTIFICATION','CLIENT_A CERTIFICATION','CLIENT_A TRAINING',
        'CLIENT_A TRAINING','CLIENT_A ','CLIENT_A SERVICE',
        'SERVICE AUDITS SALES','SERVICE, AUDITS, SALES',
        'SERVICING GOLD MEDAL','SERVICING PPGM','GOLD MEDAL TRAINING',
        'CLIENT_B CERT','CLIENT_B CGMP','CLIENT_A SERVICE',
        ]
    for kw in CLIENT_CERT_KW:
        if kw in t: return 'Client_Cert'
    # "Client A" alone in branch_techs section = Client A cert (common the portal naming)
    if area == 'branch_techs' and t.strip() in ('CLIENT_A','CLIENT_A'):
        return 'Client_Cert'

    # TECH LICENSE
    TECH_KW = ['APPLICATOR LICENSE','APPLICATORS LICENSE','APPLICATOR CARD',
        'CERTIFIED APPLICATOR','TECHNICIAN LICENSE','SERVICE TECHNICIAN LICENSE',
        'TECH LICENSE','TECH CARD','STATE LICENSE','STATE LICENCE',
        'OPERATORS LICENSE','OPERATOR LICENSE','CERTIFIED OPERATOR',
        'CERTIFIED PEST CONTROL OPERATOR','PC LICENSE','PCO',
        'PESTICIDE LICENSE','PESTICIDE APPLICATOR LICENSE',
        'PEST CONTROL LICENSE','PEST LICENSE',
        'DEPT OF AG','DEPARTMENT OF AGRICULTURE',
        'TN LICENSE','LA LICENSE','WV APPLICATOR','MDA LICENSE',
        'ARKANSAS','AR AGENTS','OK LICENSE','RHODE ISLAND LICENSE',
        'FL LIC','FL LICENSE','SP LICENSE','SP/CO LICENSE',
        'IDPH','RECERTIFICATION','REGISTRATION',
        'MANAGER LICENSE','STRUCTURAL PEST CONTROL LICENSE',
        'STATE APPLICATOR','EMPLOYEE TRAING FILE','TRAINING FILE',
        'LICENSE CARD','LICENSE HOLDER','LICENSED UNDER',
        'ID CARD','STATE ID','MISS TECH ID CARD','AGRICULTURE CARD',
        '7C','ALL APPLICATOR CARDS','RECENT ADDITION APPLICATOR',
        'PRO LICENSE','GREENPRO','GREEN PRO','QUALITY PRO',
        'PROVIDER LICENSE','APL','CERT CARD','CEUS',
        'COMMERCIAL OP','COMMERCIAL SERVICE CERTIFICATION',
        'CERTIFICATION']
    for kw in TECH_KW:
        if kw in t: return 'PC_License_Tech'
    # "License YYYY" in branch_techs = tech license
    if area == 'branch_techs' and 'LICENSE' in t:
        return 'PC_License_Tech'

    return None

# ─────────────────────────────────────────────────────────────────────
# LOAD WORKBOOK
# ─────────────────────────────────────────────────────────────────────
print(f"\nLoading: {COMPLIANCE_PATH}")
try:
    wb = openpyxl.load_workbook(COMPLIANCE_PATH)
    print("  ✓ Compliance model loaded")
except Exception as e:
    print(f"  ✗ FAILED to load compliance model: {e}")
    sys.exit(1)

# raw docs
ws_raw = wb['Raw_Document_Pull']
raw_docs_all = []
for i, row in enumerate(ws_raw.iter_rows(values_only=True)):
    if i == 0: continue
    if not row[0]: continue
    doc_area = str(row[1] or '').strip().lower()
    pull_date_raw = row[5] if len(row) > 5 else None
    raw_docs_all.append({
        'address':          norm_addr(row[0]),
        'address_raw':      row[0],
        'doc_area':         doc_area,
        'tech_name':        str(row[2] or '').strip() if doc_area != 'qa_manual' else '',
        'inspection_date':  str(row[2] or '').strip() if doc_area == 'qa_manual' else '',
        'doc_title':        str(row[3] or '').strip(),
        'exp_date':         row[4],
        'pull_date':        pull_date_raw,
    })

# For each address, only use the most recently pulled rows
# This ensures reruns overwrite old data without losing history in the tab
def _to_date(pd_raw):
    if pd_raw is None: return date.min
    if isinstance(pd_raw, datetime): return pd_raw.date()
    if isinstance(pd_raw, date): return pd_raw
    d = parse_date(pd_raw)
    return d if d else date.min

# Find latest pull_date per address
latest_pull = {}
for d in raw_docs_all:
    addr = d['address']
    pd = _to_date(d['pull_date'])
    if addr not in latest_pull or pd > latest_pull[addr]:
        latest_pull[addr] = pd

# Keep only rows from the latest pull for each address
raw_docs = [d for d in raw_docs_all
            if _to_date(d['pull_date']) == latest_pull.get(d['address'], date.min)]
print(f"  ✓ Raw docs: {len(raw_docs)} rows (most recent pull per address, of {len(raw_docs_all)} total), {len(set(d['address'] for d in raw_docs))} unique addresses")

# Last WO tech — populated from WO_Counts_Detail inside the service ticket
# block below. Initialized empty here so downstream code never KeyErrors if
# the service ticket workbook fails to load.
last_wo_tech = {}

equip_verdicts = {}

# locations
locations = {}
for sheet in ['Client_A_Locations','Client_B_Locations']:
    ws = wb[sheet]
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0: continue
        if not row[0]: continue
        addr = norm_addr(row[0])
        locations[addr] = {
            'branch':       row[3],
            'state':        str(row[4] or '').upper().strip(),
            'company':      str(row[2] or ''),
            'is_franchise': str(row[5] or 'NO').upper() == 'YES',
        }
print(f"  ✓ Locations: {len(locations)} entries")

# state exceptions
ws_st   = wb['State_Exceptions']
state_exc = {}
for i, row in enumerate(ws_st.iter_rows(values_only=True)):
    if i == 0: continue
    if not row[0]: continue
    state_exc[str(row[0]).upper().strip()] = str(row[1] or 'YES').upper().strip()
print(f"  ✓ State exceptions: {len(state_exc)} states")

# multi-drop locations (flag for manual review)
multidrop_addrs = set()
for sheet in ['Client_A_MultiDrop', 'Client_B_MultiDrop']:
    if sheet in wb.sheetnames:
        ws_md = wb[sheet]
        for i, row in enumerate(ws_md.iter_rows(values_only=True)):
            if i == 0: continue
            if not row[0]: continue
            drop_num = str(row[1]).strip() if row[1] is not None else '1'
            if drop_num in ('2', '3'):
                multidrop_addrs.add(norm_addr(row[0]))
print(f"  ✓ Multi-drop locations: {len(multidrop_addrs)} flagged for manual review")

# Map_Manual_Data — manually-entered map counts (source of truth for maps)
map_data = {}
if 'Map_Manual_Data' in wb.sheetnames:
    ws_md = wb['Map_Manual_Data']
    def _mi(v):
        if v in (None, ''): return None
        try: return int(v)
        except (ValueError, TypeError): return None
    for i, row in enumerate(ws_md.iter_rows(values_only=True)):
        if i == 0: continue
        if not row[0]: continue
        addr = norm_addr(row[0])
        map_data[addr] = {
            'erb':    _mi(row[2]),
            'irt':    _mi(row[3]),
            'ifl':    _mi(row[4]),
            'dated':  str(row[5] or '').strip().upper(),   # YES / NO / ''
            'signed': str(row[6] or '').strip().upper(),
        }
    print(f"  ✓ Map manual data: {len(map_data)} entries")

# service ticket (optional — won't crash if missing)
sow_counts     = {}
scanned_counts = {}
print(f"\nLoading service ticket report...")
try:
    wb_st = openpyxl.load_workbook(SERVICE_TICKET_PATH)
    # Month filter — service ticket Month column is stored as 1st of month (1/1/2026 etc)
    audit_m = RR['month_end'].month
    audit_y = RR['month_end'].year

    def _row_month_matches(month_val):
        """Returns True if month_val matches the audit month/year."""
        if month_val is None: return False
        if isinstance(month_val, (datetime, date)):
            d = month_val if isinstance(month_val, date) else month_val.date()
            return d.month == audit_m and d.year == audit_y
        # Try parsing string
        d = parse_date(str(month_val))
        return d is not None and d.month == audit_m and d.year == audit_y

    # Short month key used to filter WO_Counts_Detail and Equipment_Compliance
    # (those tabs store month as e.g. "Feb 2026" — match that format)
    short_audit = RR['month_end'].strftime('%b %Y')

    # SOW DEVICE COUNTS (used by map comparison; NOT the equipment verdict).
    ws_sow = wb_st['SOW_Lookup']
    def _to_int(v):
        if v in (None, ''): return 0
        try: return int(v)
        except (ValueError, TypeError): return 0   # 'CANCELLED' etc.
    for i, row in enumerate(ws_sow.iter_rows(values_only=True)):
        if i == 0: continue
        if not row[0]: continue
        addr = norm_addr(row[0])
        sow_counts[addr] = {'ERB': _to_int(row[1]),
                            'IRT': _to_int(row[5]),
                            'IFL': _to_int(row[9])}

    # TECH NAME per address — earliest non-zero WO of the audit month from
    # WO_Counts_Detail. "LAST, FIRST" gets flipped to "FIRST LAST" for matching
    # against the doc tables.
    def _flip_name(n):
        if ',' in n:
            parts = [p.strip() for p in n.split(',', 1)]
            if len(parts) == 2 and parts[1]:
                return f'{parts[1]} {parts[0]}'
        return n

    tech_candidates = {}
    ws_wod = wb_st['WO_Counts_Detail']
    for i, row in enumerate(ws_wod.iter_rows(values_only=True)):
        if i == 0: continue
        if not row[1] or not row[6]: continue
        if str(row[6]).strip() != short_audit: continue
        if (row[9] or 0) + (row[10] or 0) + (row[11] or 0) == 0:
            continue   # skip 0/0/0 birdwork/termite tickets
        addr = norm_addr(row[1])
        completed = parse_date(row[5])
        tech_candidates.setdefault(addr, []).append((completed, str(row[3] or '').strip()))
    for addr, cands in tech_candidates.items():
        cands.sort(key=lambda x: x[0] or date.max)
        last_wo_tech[addr] = _flip_name(cands[0][1])

    # EQUIPMENT VERDICT per address — pre-computed in Equipment_Compliance.
    equip_verdicts = {}
    if 'Equipment_Compliance' in wb_st.sheetnames:
        ws_eq = wb_st['Equipment_Compliance']
        for i, row in enumerate(ws_eq.iter_rows(values_only=True)):
            if i == 0: continue
            if not row[0]: continue
            if str(row[2]).strip() != short_audit: continue
            equip_verdicts[norm_addr(row[0])] = str(row[22] or 'NO_DATA').strip()
        print(f"  ✓ Equipment verdicts: {len(equip_verdicts)} for {AUDIT_MONTH}")
    else:
        print(f"  ⚠ Equipment_Compliance tab not found — run compare_equipment.py first")

    print(f"  ✓ Service ticket loaded — {len(sow_counts)} SOW rows, {len(last_wo_tech)} tech entries")
except Exception as e:
    print(f"  ⚠ Service ticket not loaded: {e}")
    print("    Equipment match will show NO_DATA (this is OK, rest will still run)")

# group docs by address
docs_by_addr = {}
for d in raw_docs:
    docs_by_addr.setdefault(d['address'], []).append(d)

# ─────────────────────────────────────────────────────────────────────
# EVALUATE EACH LOCATION
# ─────────────────────────────────────────────────────────────────────
print(f"\nEvaluating {len(docs_by_addr)} locations...")
results = []

for addr, docs in docs_by_addr.items():
    loc        = locations.get(addr)
    tech_name  = last_wo_tech.get(addr, '')
    state      = (loc or {}).get('state', '')
    branch     = (loc or {}).get('branch', '')
    company    = (loc or {}).get('company', '')
    is_fra     = (loc or {}).get('is_franchise', False)
    if branch and str(branch).startswith('9'): is_fra = True
    state_req  = state_exc.get(state, 'YES') == 'YES'

    if not loc:
        print(f"  WARN: {addr} not in Client A or Client B locations tab")

    # multi-drop locations — flag everything for manual review
    if addr in multidrop_addrs:
        original_addr = docs[0]['address_raw'] if docs else addr
        results.append({
            'audit_month':       AUDIT_MONTH,
            'address':           original_addr,
            'branch':            (loc or {}).get('branch', ''),
            'state':             (loc or {}).get('state', ''),
            'company':           (loc or {}).get('company', ''),
            'tech_name':         tech_name,
            'PC_license_branch': 'MANUAL_REVIEW',
            'PC_license_tech':   'MANUAL_REVIEW',
            'client_cert':        'MANUAL_REVIEW',
            'IPM_cert':          'MANUAL_REVIEW',
            'COI':               'MANUAL_REVIEW',
            'annual_report':     'MANUAL_REVIEW',
            'quarterly_trend':   'MANUAL_REVIEW',
            'pesticide_log':     'MANUAL_REVIEW',
            'equipment_match':   'MANUAL_REVIEW',
            'map_found':         'MANUAL_REVIEW',
            'map_erb':           '',
            'map_irt':           '',
            'map_ifl':           '',
            'map_dated':         '',
            'map_signed':        '',
            'map_vs_sow':        'MANUAL_REVIEW',
            'map_mismatch_detail': '',
            'flags':             'Multi-drop location — verify correct drop in the portal manually',
            'last_checked':      datetime.today().strftime('%m/%d/%Y'),
        })
        print(f"  ⚠ {original_addr}  MULTI-DROP — flagged for manual review")
        continue

    # starting values
    pc_branch  = 'MISSING'
    pc_tech    = 'MISSING' if state_req else 'EXEMPT'
    client_cert = 'MISSING'
    ipm_cert   = 'MISSING'
    coi        = 'MISSING'
    annual     = 'MISSING'
    trend      = 'EXEMPT' if is_fra else 'MISSING'
    pest_log   = 'EXEMPT' if is_fra else 'MISSING'
    doc_flags  = []

    for doc in docs:
        title    = doc['doc_title']
        doc_area = doc['doc_area']
        row_tech = doc['tech_name']
        es       = exp_status(doc['exp_date'])
        dtype    = classify(title, doc_area)

        # Bundled "Business & Tech License" — one PDF covering both a branch
        # license and a tech license. Must be opened, so flag both for review.
        if is_branch_tech_bundle(title):
            if pc_branch != 'PASS':
                pc_branch = 'MANUAL_REVIEW'
                doc_flags.append(f'PC_License_Branch: bundled branch+tech doc "{title}"')
            if state_req and pc_tech != 'PASS':
                pc_tech = 'MANUAL_REVIEW'
                doc_flags.append(f'PC_License_Tech: bundled branch+tech doc "{title}"')
            continue

        # Ambiguous "training cert" with no clear program name — at these
        # accounts that's almost always Client A service training, but we can't
        # be sure without opening it, so flag Client A for review (never overrides
        # a real Client A PASS found elsewhere).
        if dtype is None and 'TRAINING' in fix_typos(str(title).upper()) and 'CERT' in str(title).upper():
            if client_cert == 'MISSING':
                client_cert = 'MANUAL_REVIEW'
                doc_flags.append(f'Client_Cert: ambiguous training cert "{title}"')
            continue

        if dtype is None:
            continue

        # ── BRANCH LICENSE ───────────────────────────────────────────
        if dtype == 'PC_License_Branch':
            if   es == 'VALID'              and pc_branch != 'PASS':   pc_branch = 'PASS'
            elif es == 'MANUAL_REVIEW_DATE' and pc_branch != 'PASS':
                pc_branch = 'MANUAL_REVIEW'
                doc_flags.append(f'PC_License_Branch: date unclear "{title}" exp={doc["exp_date"]}')
            elif es == 'EXPIRED'            and pc_branch != 'PASS':
                pc_branch = 'EXPIRED'
                doc_flags.append(f'PC_License_Branch: EXPIRED "{title}"')

        # ── COI ──────────────────────────────────────────────────────
        # NOTE: the portal expiration dates are unreliable (~15% wrong).
        # Rule: if the title clearly contains the audit year (e.g. "Insurance 2026"),
        # that overrides the expiration date and counts as PASS.
        # If no year in title, we still check the expiration date as a fallback.
        elif dtype == 'COI':
            if has_year(title, RR['annual_year']):
                # Title has 2026 → trust the title, not the date
                coi = 'PASS'
            elif has_year(title, RR['annual_year'] - 1):
                # Title clearly says last year (e.g. "2025 COI") → wrong year
                if coi not in ('PASS',):
                    coi = 'WRONG_YEAR'
                doc_flags.append(f'COI: title shows {RR["annual_year"]-1}, need {RR["annual_year"]} "{title}"')
            else:
                # No year in title — fall back to the expiration date.
                # A COI with a valid (not-yet-expired) date is active coverage,
                # so it PASSES even if the exp year is past the audit year
                # (e.g. expires 1/1/2027 — still covers a Feb 2026 audit).
                if es == 'VALID':
                    coi = 'PASS'
                elif es == 'EXPIRED' and coi != 'PASS':
                    coi = 'EXPIRED'
                    doc_flags.append(f'COI: no year in title and EXPIRED "{title}" exp={doc["exp_date"]}')
                elif es == 'MANUAL_REVIEW_DATE' and coi != 'PASS':
                    coi = 'MANUAL_REVIEW'
                    doc_flags.append(f'COI: no year in title, date unclear "{title}" exp={doc["exp_date"]}')

        # ── ANNUAL REPORT ────────────────────────────────────────────
        elif dtype == 'Annual_Report':
            if doc_area != 'qa_manual': continue
            if has_year(title, RR['annual_year']):
                annual = 'PASS'
            elif annual != 'PASS':
                # Check inspection_date year as fallback
                insp = parse_inspection_date(doc.get('inspection_date',''), RR['annual_year'])
                if insp and insp.year == RR['annual_year']:
                    annual = 'PASS'  # inspection date confirms current year
                elif has_year(title, RR['annual_year'] - 1):
                    # Explicitly prior year — MISSING not MANUAL_REVIEW
                    doc_flags.append(f'Annual_Report: {RR["annual_year"]-1} doc found, need {RR["annual_year"]} "{title}"')
                elif insp and insp.year < RR['annual_year']:
                    # Inspection date shows prior year — MISSING
                    doc_flags.append(f'Annual_Report: inspection date {insp.year}, need {RR["annual_year"]} "{title}"')
                else:
                    annual = 'MANUAL_REVIEW'
                    doc_flags.append(f'Annual_Report: year unclear, need {RR["annual_year"]} "{title}"')

        # ── QUARTERLY TREND ──────────────────────────────────────────
        elif dtype == 'Quarterly_Trend':
            if doc_area != 'qa_manual' or trend == 'EXEMPT': continue
            if has_target_quarter(title):
                trend = 'PASS'
            elif has_explicit_wrong_year(title):
                doc_flags.append(f'Quarterly_Trend: wrong year, need Q{RR["qn"]} {RR["qy"]} "{title}"')
            elif trend != 'PASS':
                # Try inspection_date year hint for undated docs
                insp = parse_inspection_date(doc.get('inspection_date',''), RR['annual_year'])
                iqy = implied_quarter_year(title, insp)
                if iqy and iqy == (RR['qn'], RR['qy']):
                    trend = 'PASS'  # inspection date confirms correct quarter/year
                elif iqy and iqy != (RR['qn'], RR['qy']):
                    doc_flags.append(f'Quarterly_Trend: inferred Q{iqy[0]} {iqy[1]}, need Q{RR["qn"]} {RR["qy"]} "{title}"')
                else:
                    trend = 'MANUAL_REVIEW'
                    doc_flags.append(f'Quarterly_Trend: need Q{RR["qn"]} {RR["qy"]} "{title}"')

        # ── PESTICIDE LOG ────────────────────────────────────────────
        elif dtype == 'Pesticide_Log':
            if doc_area != 'qa_manual' or pest_log == 'EXEMPT': continue
            if has_target_quarter(title):
                pest_log = 'PASS'
            elif has_explicit_wrong_year(title):
                doc_flags.append(f'Pesticide_Log: wrong year, need Q{RR["qn"]} {RR["qy"]} "{title}"')
            elif pest_log != 'PASS':
                # Try inspection_date year hint for undated docs
                insp = parse_inspection_date(doc.get('inspection_date',''), RR['annual_year'])
                iqy = implied_quarter_year(title, insp)
                if iqy and iqy == (RR['qn'], RR['qy']):
                    pest_log = 'PASS'
                elif iqy and iqy != (RR['qn'], RR['qy']):
                    doc_flags.append(f'Pesticide_Log: inferred Q{iqy[0]} {iqy[1]}, need Q{RR["qn"]} {RR["qy"]} "{title}"')
                else:
                    pest_log = 'MANUAL_REVIEW'
                    doc_flags.append(f'Pesticide_Log: need Q{RR["qn"]} {RR["qy"]} "{title}"')

        # ── TECH-LEVEL DOCS ──────────────────────────────────────────
        # Principle: if a doc of this type EXISTS but we can't confirm whose
        # it is (no tech-name match) or which year, the column becomes
        # MANUAL_REVIEW rather than MISSING. A confirmed match can still PASS,
        # and we never downgrade a PASS already found at this location.
        elif dtype in ('PC_License_Tech', 'Client_Cert', 'IPM_Cert'):
            if doc_area == 'qa_manual': continue
            # Agents/all-tech cards in insurance section count as branch-level tech license
            if dtype == 'PC_License_Tech' and doc_area == 'insurance':
                if es == 'VALID' and pc_tech != 'PASS':
                    pc_tech = 'PASS'
                continue
            name_match = bool(tech_name) and (
                row_tech_matches(row_tech, tech_name) or tech_in_title(tech_name, title))

            if dtype == 'PC_License_Tech' and state_req:
                if name_match:
                    if   es == 'VALID':              pc_tech = 'PASS'
                    elif es == 'MANUAL_REVIEW_DATE' and pc_tech != 'PASS':
                        pc_tech = 'MANUAL_REVIEW'
                        doc_flags.append(f'PC_License_Tech: date unclear for {tech_name} "{title}" exp={doc["exp_date"]}')
                    elif es == 'EXPIRED' and pc_tech != 'PASS':
                        pc_tech = 'EXPIRED'
                        doc_flags.append(f'PC_License_Tech: EXPIRED for {tech_name} "{title}"')
                elif pc_tech == 'MISSING':
                    pc_tech = 'MANUAL_REVIEW'
                    doc_flags.append(f'PC_License_Tech: license doc found, tech unconfirmed "{title}"')

            elif dtype == 'Client_Cert':
                if name_match:
                    client_cert = 'PASS'
                elif client_cert == 'MISSING':
                    client_cert = 'MANUAL_REVIEW'
                    doc_flags.append(f'Client_Cert: cert found, tech unconfirmed "{title}"')

            elif dtype == 'IPM_Cert':
                if name_match:
                    if has_year(title, RR['annual_year']):
                        ipm_cert = 'PASS'
                    elif has_year(title, RR['annual_year'] - 1):
                        if ipm_cert != 'PASS': ipm_cert = 'WRONG_YEAR'
                        doc_flags.append(f'IPM_Cert: 2025 found, need 2026 "{title}"')
                    else:
                        if ipm_cert != 'PASS': ipm_cert = 'MANUAL_REVIEW'
                        doc_flags.append(f'IPM_Cert: year unclear "{title}"')
                elif ipm_cert == 'MISSING':
                    ipm_cert = 'MANUAL_REVIEW'
                    doc_flags.append(f'IPM_Cert: doc found, tech/year unconfirmed "{title}"')

    # ── EQUIPMENT MATCH (from Equipment_Compliance tab) ──────────────
    # Single source of truth — verdict pre-computed by compare_equipment.py.
    equip = equip_verdicts.get(addr, 'NO_DATA')

    # ── MAP COMPARISON (Map_Manual_Data vs SOW device counts) ────────
    md = map_data.get(addr)
    sc = sow_counts.get(addr)
    if md is None:
        map_found_val = 'NO'
        map_erb = map_irt = map_ifl = ''
        map_dated_val = map_signed_val = ''
        map_vs_sow = 'NO_MAP'
        map_mismatch = ''
    else:
        map_erb, map_irt, map_ifl = md['erb'], md['irt'], md['ifl']
        map_dated_val, map_signed_val = md['dated'], md['signed']
        map_found_val = 'YES' if any(v is not None for v in (map_erb, map_irt, map_ifl)) else 'NO'
        if not sc:
            map_vs_sow = 'NO_SOW'; map_mismatch = ''
        elif map_erb is None and map_irt is None and map_ifl is None:
            map_vs_sow = 'NO_MAP_COUNTS'; map_mismatch = ''
        else:
            issues = []
            if map_erb is not None and map_erb != sc['ERB']:
                issues.append(f'ERB: Map={map_erb} SOW={sc["ERB"]}')
            if map_irt is not None and map_irt != sc['IRT']:
                issues.append(f'IRT: Map={map_irt} SOW={sc["IRT"]}')
            if map_ifl is not None and map_ifl != sc['IFL']:
                issues.append(f'IFL: Map={map_ifl} SOW={sc["IFL"]}')
            map_vs_sow = 'MISMATCH' if issues else 'MATCH'
            map_mismatch = ' | '.join(issues)

    # ── COMPLIANCE % ─────────────────────────────────────────────────
    checks = {
        'PC_License_Branch': pc_branch, 'PC_License_Tech': pc_tech,
        'Client_Cert': client_cert,        'IPM_Cert': ipm_cert,
        'COI': coi,                      'Annual_Report': annual,
        'Quarterly_Trend': trend,        'Pesticide_Log': pest_log,
    }
    applicable     = [k for k,v in checks.items() if v != 'EXEMPT']
    passed         = [k for k in applicable if checks[k] == 'PASS']
    checks_app     = len(applicable)
    checks_pass    = len(passed)
    pct            = round(checks_pass / checks_app * 100, 1) if checks_app else 0

    # build flags string
    flag_issues = [f'{k}: {v}' for k,v in checks.items() if v in ('MISSING','EXPIRED','WRONG_YEAR','MANUAL_REVIEW')]
    if equip not in ('PASS', 'CANCELLED', 'NO_DATA', ''):
        flag_issues.append(f'EQUIPMENT: {equip}')
    if map_vs_sow == 'MISMATCH':
        flag_issues.append(f'MAP MISMATCH: {map_mismatch}')
    flag_issues.extend(doc_flags)
    flags_str = ' | '.join(flag_issues) if flag_issues else 'NONE'

    original_addr = docs[0]['address_raw'] if docs else addr

    results.append({
        'audit_month':       AUDIT_MONTH,
        'address':           original_addr,
        'branch':            branch,
        'state':             state,
        'company':           company,
        'tech_name':         tech_name,
        'PC_license_branch': pc_branch,
        'PC_license_tech':   pc_tech,
        'client_cert':        client_cert,
        'IPM_cert':          ipm_cert,
        'COI':               coi,
        'annual_report':     annual,
        'quarterly_trend':   trend,
        'pesticide_log':     pest_log,
        'equipment_match':   equip,
        'map_found':         map_found_val,
        'map_erb':           map_erb if map_erb is not None else '',
        'map_irt':           map_irt if map_irt is not None else '',
        'map_ifl':           map_ifl if map_ifl is not None else '',
        'map_dated':         map_dated_val,
        'map_signed':        map_signed_val,
        'map_vs_sow':        map_vs_sow,
        'map_mismatch_detail': map_mismatch,
        'flags':             flags_str,
        'last_checked':      datetime.today().strftime('%m/%d/%Y'),
    })
    print(f"  ✓ {original_addr}  {pct}%  tech=[{tech_name}]  branch_lic={pc_branch}  tech_lic={pc_tech}  client={client_cert}  ipm={ipm_cert}  coi={coi}  annual={annual}  trend={trend}  log={pest_log}")

# ─────────────────────────────────────────────────────────────────────
# WRITE TO COMPLIANCE_EVAL
# ─────────────────────────────────────────────────────────────────────
HEADERS = [
    'audit_month','address','branch','state','company','tech_name',
    'PC_license_branch','PC_license_tech','client_cert','IPM_cert',
    'COI','annual_report','quarterly_trend','pesticide_log',
    'equipment_match',
    'map_found','map_erb','map_irt','map_ifl',
    'map_dated','map_signed','map_vs_sow','map_mismatch_detail',
    'flags','last_checked'
]

print(f"\nWriting {len(results)} rows to Compliance_Eval...")
ws_eval = wb['Compliance_Eval']

# Always sync the header row to the current HEADERS list. Old data rows below
# stay put. If the previous header set was different, OLD ROWS' columns past
# the changed point may be misaligned vs. the new headers — delete pre-rewire
# rows if you want a clean Power BI view.
existing_headers = [ws_eval.cell(row=1, column=c).value for c in range(1, len(HEADERS) + 1)]
if existing_headers != HEADERS:
    if any(existing_headers):
        print("  ⚠ Header row changed — old data rows may be misaligned in the new columns.")
        print("    Recommended: delete pre-rewire rows once Power BI is wired up.")
    for col, h in enumerate(HEADERS, 1):
        ws_eval.cell(row=1, column=col, value=h)

# Build index of existing rows keyed by (address, audit_month) -> row number
existing_index = {}
for row_idx in range(2, ws_eval.max_row + 1):
    ex_addr  = ws_eval.cell(row=row_idx, column=HEADERS.index('address') + 1).value
    ex_month = ws_eval.cell(row=row_idx, column=HEADERS.index('audit_month') + 1).value
    if ex_addr and ex_month:
        existing_index[(str(ex_addr).strip(), str(ex_month).strip())] = row_idx

# Overwrite matching rows; append new ones
overwritten = 0
appended = 0
next_row = ws_eval.max_row + 1 if ws_eval.max_row > 1 else 2

for result in results:
    key = (str(result.get('address', '')).strip(), str(result.get('audit_month', '')).strip())
    if key in existing_index:
        row_idx = existing_index[key]
        overwritten += 1
    else:
        row_idx = next_row
        next_row += 1
        appended += 1
    for col_idx, h in enumerate(HEADERS, 1):
        ws_eval.cell(row=row_idx, column=col_idx, value=result.get(h, ''))

wb.save(COMPLIANCE_PATH)
print(f"✓ Saved. {overwritten} rows overwritten, {appended} new rows appended to Compliance_Eval.")

# ─────────────────────────────────────────────────────────────────────
# WRITE / UPDATE Map_Manual_Data TAB
# Only writes address and map_found — never overwrites user-filled data
# ─────────────────────────────────────────────────────────────────────
MAP_MANUAL_HEADERS = [
    'address','map_found','map_erb','map_irt','map_ifl',
    'map_dated','map_signed','map_status','last_verified','map_flags'
]
MAP_AUTO_COLS = {'address', 'map_found'}  # only these are auto-written

wb2 = openpyxl.load_workbook(COMPLIANCE_PATH)
if 'Map_Manual_Data' not in wb2.sheetnames:
    ws_map = wb2.create_sheet('Map_Manual_Data')
    for col, h in enumerate(MAP_MANUAL_HEADERS, 1):
        ws_map.cell(row=1, column=col, value=h)
    print("  ✓ Map_Manual_Data tab created")
else:
    ws_map = wb2['Map_Manual_Data']

# Build index of existing addresses in Map_Manual_Data
existing_rows = {}
for i, row in enumerate(ws_map.iter_rows(min_row=2, values_only=False), 2):
    addr_cell = row[0].value
    if addr_cell:
        existing_rows[norm_addr(str(addr_cell))] = i

for result in results:
    addr_norm = norm_addr(str(result.get('address','')))
    map_found_val = result.get('map_found', 'MISSING')
    if addr_norm in existing_rows:
        # Update map_found only (col B = index 2)
        row_idx = existing_rows[addr_norm]
        ws_map.cell(row=row_idx, column=2, value=map_found_val)
    else:
        # New address — add row with address + map_found, leave rest blank
        next_row = ws_map.max_row + 1
        ws_map.cell(row=next_row, column=1, value=result.get('address',''))
        ws_map.cell(row=next_row, column=2, value=map_found_val)

wb2.save(COMPLIANCE_PATH)
print(f"✓ Map_Manual_Data updated — {len(results)} addresses")
print("\nDone!")

# ─────────────────────────────────────────────────────────────────────
# FILL Doc_Compliance.Technician IN THE SERVICE TICKET WORKBOOK
#
# JOIN KEY: (normalized_address, month_word)
#   audit_month "July 2026" -> month_word "July"
# Only touches rows where the target tab's Month column == month_word,
# so a July run never disturbs a June row and vice versa.
#
# Rules:
#   - Only fills BLANK Technician cells by default (protects manual typing).
#   - Only writes when Compliance_Eval actually has a tech for that row.
#   - Flip OVERWRITE_EXISTING_TECH to True if you ever want to force-refresh.
# ─────────────────────────────────────────────────────────────────────
DOC_TAB                 = 'Doc_Compliance_June'   # yes, misnamed — it holds both months
OVERWRITE_EXISTING_TECH = False

month_word = AUDIT_MONTH.split()[0]  # "July 2026" -> "July"

# 1) Build the lookup dictionary from THIS run's results.
#    Key: (normalized_address, month_word)   Value: tech name
addr_month_to_tech = {}
for r in results:
    if r.get('tech_name'):
        addr_month_to_tech[(norm_addr(r['address']), month_word)] = r['tech_name']

print(f"\nFilling Technician column in '{DOC_TAB}' for month '{month_word}'...")
wb_st = openpyxl.load_workbook(SERVICE_TICKET_PATH)

if DOC_TAB not in wb_st.sheetnames:
    print(f"  ⚠ Tab '{DOC_TAB}' not found in {SERVICE_TICKET_PATH}. Skipping.")
else:
    ws_doc = wb_st[DOC_TAB]

    # 2) Find columns by header name (survives reordering).
    header_row = next(ws_doc.iter_rows(min_row=1, max_row=1, values_only=False))
    header_map = {str(c.value).strip().lower(): c.column for c in header_row if c.value}
    addr_col  = header_map.get('address')
    month_col = header_map.get('month')
    tech_col  = header_map.get('technician')

    if not all([addr_col, month_col, tech_col]):
        print(f"  ⚠ '{DOC_TAB}' missing Address, Month, or Technician header. Skipping.")
    else:
        filled = already = no_match = other_month = 0
        for row_idx in range(2, ws_doc.max_row + 1):
            addr_val  = ws_doc.cell(row=row_idx, column=addr_col).value
            row_month = ws_doc.cell(row=row_idx, column=month_col).value
            if not addr_val or not row_month:
                continue
            if str(row_month).strip() != month_word:
                other_month += 1        # e.g. a June row during a July run
                continue
            src_tech = addr_month_to_tech.get((norm_addr(str(addr_val)), month_word))
            if not src_tech:
                no_match += 1
                continue
            current = ws_doc.cell(row=row_idx, column=tech_col).value
            if current and str(current).strip() and not OVERWRITE_EXISTING_TECH:
                already += 1
                continue
            ws_doc.cell(row=row_idx, column=tech_col, value=src_tech)
            filled += 1

        wb_st.save(SERVICE_TICKET_PATH)
        print(f"  ✓ Filled {filled}  |  already had tech: {already}  "
              f"|  no source match: {no_match}  |  other-month rows left alone: {other_month}")
