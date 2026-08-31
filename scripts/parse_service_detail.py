import openpyxl
import os
import re
import glob
from collections import defaultdict

# ─────────────────────────────────────────────────────────────────────
# NOTE ON GENERICIZATION
# Client, employer, and system names have been replaced throughout this
# file for public release: CLIENT_A / CLIENT_B are two national accounts,
# and "the portal" is a proprietary service-management platform.
#
# The document-title keyword lists below are therefore illustrative
# rather than runnable — in production they contain the real strings that
# appear in scanned document titles. The matching logic, thresholds, and
# control flow are unchanged.
# ─────────────────────────────────────────────────────────────────────

# ── HOW TO RUN ────────────────────────────────────────────────────────────────
# 1. Export the "Service Detail Report by Workorder" from the portal (Client A + Client B)
# 2. Save in: C:\Automation\Raw Data\
#    File names can be anything as long as they start with:
#      ClientA-_Raw_Data_   (and end in .xlsx)
#      ClientB-_Raw_Data_   (and end in .xlsx)
#    Examples that all work:
#      ClientA-_Raw_Data_Jan1_to_May_31st_svc_detail_by_WO.xlsx
#      ClientA-_Raw_Data_2026-06-01_to_2026-06-15_svc_detail_by_WO.xlsx
#      ClientA-_Raw_Data_LATEST.xlsx
# 3. Open Command Prompt
# 4. Type:  py C:\Automation\parse_service_detail.py
# 5. Press Enter
#
# OVERLAPPING DATE RANGES ARE SAFE:
#   If the same work order appears in two files (e.g. Jan-May archive AND a
#   newer pull that also includes some May dates), it's counted ONCE.
#   Dedupe is by WORK ORDER number.
#
# BEFORE RUNNING:
#   Make sure Service_Ticket_Report.xlsx is CLOSED in Excel.
#
# WHAT IT OUTPUTS:
#   Tab WO_Counts_Detail in Service_Ticket_Report.xlsx.
#   One row per UNIQUE work order across all raw files.
# ─────────────────────────────────────────────────────────────────────────────

RAW_DATA_FOLDER = r"C:\Automation\Raw Data"
OUTPUT_FOLDER   = r"C:\Automation"
PEPSI_PATTERN   = "ClientA-_Raw_Data_*.xlsx"
FRITO_PATTERN   = "ClientB-_Raw_Data_*.xlsx"
TICKET_FILE     = "Service_Ticket_Report.xlsx"
OUTPUT_TAB      = "WO_Counts_Detail"

# ── WHICH PROGRAM TYPES COUNT ─────────────────────────────────────────────────
# Only real recurring equipment service. PC 1st Service handled specially below.
# Everything else (Odd Job, Follow Up, Callback, PPGM, Quality Inspection,
# Materials, Tech Phone Call, Bait Monitoring) is ignored automatically.
KEEP_ACTIVITIES = {'PC Standard'}

# ── TRAP TYPE CATEGORIES ──────────────────────────────────────────────────────
ERB_TYPES = {'E. Rodent Bait Station'}
IRT_TYPES = {'Rodent Traps'}
IFL_TYPES = {'Insect Light Traps', 'Fly Bait Station'}  # Fly Bait Station confirmed = fly light

# Decided by ZONE (interior vs exterior) — the name alone is unreliable:
ZONE_BASED_TYPES = {'I. Rodent Bait Station', 'E. Rodent Traps', 'Glue Board'}

# Flagged for manual review (genuinely ambiguous). Zone-based types flag themselves.
REVIEW_TYPES = set()

# Completely ignored — not countable ERB/IRT/IFL equipment
IGNORE_TYPES = {'Inspection', 'Pheromone Traps', 'Accessories',
                'Insect Glue Board', 'Insect Bait Station'}

# Zone keyword hints used by the zone-based types above.
ZONE_INTERIOR_HINTS = ('TINCAT', 'TIN CAT', 'TIN-CAT', 'INTERIOR', 'RECEIVING',
                       'WAREHOUSE', 'INTERIOR TRAPS')
ZONE_EXTERIOR_HINTS = ('EXTERIOR', 'OUTSIDE', 'PERIMETER')
# 'EXT' as a standalone word/abbreviation (not inside words like 'Next' or 'Textile')
# Insect/monitor zones — a Glue Board here is an insect monitor, not a rodent device
ZONE_INSECT_HINTS   = ('INSECT MONITOR', 'PEST MONITOR', 'BREAKROOM',
                       'INSECT TREATMENT', 'INSECT MONITORS')

COUNT_STATUSES = {'Serviced', 'Replaced', 'Added'}

MONTH_NAMES = {1:'Jan',2:'Feb',3:'Mar',4:'Apr',5:'May',6:'Jun',
               7:'Jul',8:'Aug',9:'Sep',10:'Oct',11:'Nov',12:'Dec'}

# ── HELPERS ───────────────────────────────────────────────────────────────────
def clean_address(raw):
    """Strip location-type prefix and lot info, keep just the street address."""
    if not raw:
        return ''
    s = str(raw).strip().upper()
    s = re.sub(r'\s+LOT\s*#.*$', '', s, flags=re.IGNORECASE).strip()
    s = re.sub(r'^(CLIENT_A|CLIENT_B)[\s\-/]+[A-Z\s]+?\s+(?=\d)', '', s).strip()
    if ',' in s:
        s = s.split(',')[0].strip()
    return s

def find_label(row, label):
    """Find a label cell anywhere in the row; return its column index or None.
    The report shifts header labels between columns, so we never assume a
    fixed column — we locate the label text, then read the value at an offset."""
    for j, v in enumerate(row):
        if v and label in str(v):
            return j
    return None

def categorize_by_zone(trap, zone):
    """For zone-based trap types, decide ERB / IRT / IGNORE / REVIEW from the zone.
       - I. Rodent Bait Station & E. Rodent Traps: interior zone -> IRT, exterior -> ERB
       - Glue Board: insect/monitor zone -> IGNORE, interior rodent zone -> IRT, else REVIEW
       Anything genuinely unclear returns REVIEW so it is flagged, never guessed.
       Note: an explicit 'EXTERIOR'/'EXT' wins even if an interior word is also present
       (e.g. 'Warehouse Exterior' is exterior), since the outdoor designation is specific."""
    z = str(zone or '').upper()
    is_int    = any(h in z for h in ZONE_INTERIOR_HINTS)
    # exterior if a full exterior word appears, OR 'EXT' as a standalone token
    is_ext    = any(h in z for h in ZONE_EXTERIOR_HINTS) or bool(re.search(r'\bEXT\b', z))
    is_insect = any(h in z for h in ZONE_INSECT_HINTS)

    if trap == 'Glue Board':
        if is_insect:
            return 'IGNORE'
        if is_int and not is_ext:
            return 'IRT'
        return 'REVIEW'

    # I. Rodent Bait Station and E. Rodent Traps
    if is_ext:           # explicit exterior wins (handles 'Warehouse Exterior', 'Ext')
        return 'ERB'
    if is_int:
        return 'IRT'
    return 'REVIEW'

# ── PARSE ONE FILE ────────────────────────────────────────────────────────────
def parse_file(path, client_label):
    print(f"\nOpening {client_label} file: {os.path.basename(path)}")
    print("  (Large file with many tabs — may take a moment...)")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    # First pass: which addresses ever have a PC Standard service?
    # (used to tell "new" vs "existing" for PC 1st Service)
    standard_addrs = set()
    for sheet_name in wb.sheetnames:
        if sheet_name == 'P-About':
            continue
        ws = wb[sheet_name]
        address = activity = None
        for row in ws.iter_rows(values_only=True):
            if not row:
                continue
            c2 = str(row[2] or '').strip() if len(row) > 2 else ''
            if 'Location:' in c2 and len(row) > 4 and row[4]:
                address = clean_address(row[4])
            la = find_label(row, 'Service Activity')
            if la is not None and len(row) > la + 2 and row[la + 2]:
                activity = str(row[la + 2]).strip()
            if c2 == 'Station Inspection Scan Timestamp':
                break
        if address and activity == 'PC Standard':
            standard_addrs.add(address)

    # Second pass: extract every qualifying work order
    rows_out = []
    new_addr_flags = set()

    for sheet_name in wb.sheetnames:
        if sheet_name == 'P-About':
            continue
        ws = wb[sheet_name]

        raw_address = ticket = tech = branch = completed = activity = None
        data_started = False
        erb = irt = ifl = 0
        review_found = set()
        # Track unique devices per category so a station scanned twice in the
        # same work order only counts once (dedup by barcode, fallback to name).
        seen_erb = set()
        seen_irt = set()
        seen_ifl = set()

        for row in ws.iter_rows(values_only=True):
            if not row:
                continue
            c2 = str(row[2] or '').strip() if len(row) > 2 else ''

            # Header fields — located by label text, value read at offset
            if 'Location:' in c2 and len(row) > 4 and row[4]:
                raw_address = str(row[4]).strip()
            if 'Ticket Number' in c2 and len(row) > 4 and row[4]:
                ticket = row[4]
            la = find_label(row, 'Service Activity')
            if la is not None and len(row) > la + 2 and row[la + 2]:
                activity = str(row[la + 2]).strip()
            lb = find_label(row, 'Service Branch')
            if lb is not None and len(row) > lb + 3 and row[lb + 3]:
                branch = str(row[lb + 3]).strip()
            lt = find_label(row, 'Service Tech')
            if lt is not None and len(row) > lt + 3 and row[lt + 3]:
                tech = str(row[lt + 3]).strip()
            lc = find_label(row, 'Completed:')
            if lc is not None and len(row) > lc + 3 and row[lc + 3]:
                completed = row[lc + 3]

            if c2 == 'Station Inspection Scan Timestamp':
                data_started = True
                continue
            if not data_started or len(row) < 14:
                continue

            trap   = str(row[6]  or '').strip()
            status = str(row[13] or '').strip()
            zone   = str(row[3]  or '').strip()
            name   = row[9]  if len(row) > 9  else ''
            barcode = row[11] if len(row) > 11 else ''

            if not trap or status not in COUNT_STATUSES:
                continue

            # Unique key for this physical device — barcode if present, else name.
            # Prefixed with trap so the same barcode under two trap types stays distinct.
            dev_key = f"{trap}|{barcode if barcode else name}"

            if trap in IGNORE_TYPES:
                continue
            elif trap in ERB_TYPES:
                if dev_key not in seen_erb:
                    seen_erb.add(dev_key)
                    erb += 1
            elif trap in IRT_TYPES:
                if dev_key not in seen_irt:
                    seen_irt.add(dev_key)
                    irt += 1
            elif trap in IFL_TYPES:
                if dev_key not in seen_ifl:
                    seen_ifl.add(dev_key)
                    ifl += 1
            elif trap in ZONE_BASED_TYPES:
                cat = categorize_by_zone(trap, zone)
                if cat == 'IRT':
                    if dev_key not in seen_irt:
                        seen_irt.add(dev_key)
                        irt += 1
                elif cat == 'ERB':
                    if dev_key not in seen_erb:
                        seen_erb.add(dev_key)
                        erb += 1
                elif cat == 'IGNORE':
                    pass
                else:  # REVIEW
                    review_found.add(f'{trap} (ambiguous zone: {zone})')
            elif trap in REVIEW_TYPES:
                review_found.add(trap)
            else:
                review_found.add(f'UNKNOWN: {trap}')

        # ── Decide whether to keep this work order ────────────────────────
        if not (ticket and raw_address):
            continue

        addr_key = clean_address(raw_address)

        keep = False
        if activity in KEEP_ACTIVITIES:
            keep = True
        elif activity == 'PC 1st Service':
            # Existing address → ignore (add-on). New address → keep + flag.
            if addr_key not in standard_addrs:
                keep = True
                new_addr_flags.add(addr_key)
                review_found.add('NEW ADDRESS — needs SOW setup')

        if not keep:
            continue

        if completed and hasattr(completed, 'month'):
            month_label   = f"{MONTH_NAMES[completed.month]} {completed.year}"
            completed_str = completed.strftime('%m/%d/%Y')
        else:
            month_label   = 'UNKNOWN'
            completed_str = ''

        rows_out.append({
            'work_order':   ticket,
            'address':      addr_key,
            'raw_address':  raw_address,
            'technician':   tech or '',
            'branch':       branch or '',
            'completed':    completed_str,
            'month':        month_label,
            'client':       client_label,
            'activity':     activity,
            'erb':          erb,
            'irt':          irt,
            'ifl':          ifl,
            'review_flags': ', '.join(sorted(review_found)) if review_found else '',
        })

    wb.close()
    print(f"  {client_label}: {len(rows_out)} work orders kept | "
          f"{len(new_addr_flags)} new addresses flagged")
    return rows_out

# ── RUN ALL MATCHING FILES ────────────────────────────────────────────────────
print("=== Parse Service Detail Reports (Client A + Client B) ===")
print(f"Raw data folder: {RAW_DATA_FOLDER}")

pepsi_files = sorted(glob.glob(os.path.join(RAW_DATA_FOLDER, PEPSI_PATTERN)))
frito_files = sorted(glob.glob(os.path.join(RAW_DATA_FOLDER, FRITO_PATTERN)))

if not pepsi_files and not frito_files:
    print(f"\n  ERROR: No raw data files found in {RAW_DATA_FOLDER}")
    print(f"  Expected files starting with 'ClientA-_Raw_Data_' or 'ClientB-_Raw_Data_'")
    raise SystemExit(1)

print(f"\nFound {len(pepsi_files)} Client A file(s) and {len(frito_files)} Client B file(s):")
for f in pepsi_files + frito_files:
    print(f"   • {os.path.basename(f)}")

all_rows = []
for f in pepsi_files:
    all_rows += parse_file(f, 'CLIENT_A')
for f in frito_files:
    all_rows += parse_file(f, 'CLIENT_B')

# ── DEDUPE by work order ────────────────────────────────────────────────────
# A WO appearing in two files (overlapping date ranges) keeps the first one.
# Data should be identical since both come from the same the portal source.
before = len(all_rows)
seen = set()
deduped = []
for r in all_rows:
    wo = r['work_order']
    if wo in seen: continue
    seen.add(wo)
    deduped.append(r)
all_rows = deduped
print(f"\nDeduplication: {before} rows -> {len(all_rows)} unique work orders "
      f"({before - len(all_rows)} duplicates removed)")

flagged = [r for r in all_rows if r['review_flags']]
if flagged:
    print(f"\nWork orders with review flags: {len(flagged)}")
    flag_types = defaultdict(int)
    for r in flagged:
        for f in r['review_flags'].split(', '):
            flag_types[f] += 1
    for f, c in sorted(flag_types.items(), key=lambda x: -x[1]):
        print(f"   {c:4} x  {f}")

# ── WRITE TO TICKET REPORT ────────────────────────────────────────────────────
ticket_path = os.path.join(OUTPUT_FOLDER, TICKET_FILE)
print(f"\nOpening Service Ticket Report...")
wb_ticket = openpyxl.load_workbook(ticket_path)

if OUTPUT_TAB in wb_ticket.sheetnames:
    del wb_ticket[OUTPUT_TAB]
    print(f"  Removed old {OUTPUT_TAB} tab")

ws_out = wb_ticket.create_sheet(OUTPUT_TAB)

headers = [
    'WORK ORDER', 'ADDRESS', 'RAW_ADDRESS', 'TECHNICIAN', 'BRANCH',
    'COMPLETED', 'MONTH', 'CLIENT', 'ACTIVITY',
    'ERB', 'IRT', 'IFL', 'REVIEW_FLAGS',
]
ws_out.append(headers)

for r in sorted(all_rows, key=lambda x: (x['client'], x['address'], x['completed'])):
    ws_out.append([
        r['work_order'], r['address'], r['raw_address'], r['technician'],
        r['branch'], r['completed'], r['month'], r['client'], r['activity'],
        r['erb'], r['irt'], r['ifl'], r['review_flags'],
    ])

wb_ticket.save(ticket_path)

print(f"\n=== Done! ===")
print(f"  {len(all_rows)} work order rows written to '{OUTPUT_TAB}' tab")
print(f"  Saved to: {ticket_path}")
print()
print("Next steps:")
print(f"  1. Open {TICKET_FILE}")
print(f"  2. Go to the {OUTPUT_TAB} tab")
print("  3. Review rows with REVIEW_FLAGS — especially NEW ADDRESS ones")
print("  4. Spot-check a few addresses against your manual counts")
