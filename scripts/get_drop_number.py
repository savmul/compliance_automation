import sys
import re
import openpyxl

EXCEL_PATH         = r"C:\Automation\portal_Compliance_Model.xlsx"
TEMP_ADDRESS_PATH  = r"C:\Automation\temp_address.txt"
TEMP_DROPNUMBER_PATH = r"C:\Automation\temp_dropnumber.txt"

def normalize_address(value: str) -> str:
    value = str(value or "").upper().strip()
    value = re.sub(r"[.,]", "", value)
    value = re.sub(r"\s+", " ", value)
    return value.strip()

def main():
    brand = "client_a"
    if len(sys.argv) > 1:
        brand = str(sys.argv[1]).upper().strip()

    sheet_name = "client_b_MultiDrop" if brand == "client_b" else "cleint_a_MultiDrop"

    # Try multiple encodings for temp_address.txt
    target_address = None
    for enc in ('utf-16', 'utf-8-sig', 'utf-8', 'latin-1'):
        try:
            with open(TEMP_ADDRESS_PATH, "r", encoding=enc) as f:
                target_address = normalize_address(f.read())
            if target_address:
                break
        except Exception:
            continue

    if not target_address:
        print("ERROR: could not read temp_address.txt")
        raise SystemExit(1)

    print(f"DEBUG: looking up address: {target_address}")

    try:
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    except FileNotFoundError:
        print(f"ERROR: workbook not found: {EXCEL_PATH}")
        raise SystemExit(1)

    if sheet_name not in wb.sheetnames:
        print(f"ERROR: sheet not found: {sheet_name}")
        raise SystemExit(1)

    ws = wb[sheet_name]
    drop_number = "1"

    for row in ws.iter_rows(min_row=2, values_only=True):
        row_address = normalize_address(row[0])
        row_drop = str(row[1]).strip() if row[1] is not None else ""
        if not row_address:
            continue
        if row_address == target_address and row_drop in {"2", "3"}:
            drop_number = row_drop
            print(f"DEBUG: matched {row_address} → drop {drop_number}")
            break

    with open(TEMP_DROPNUMBER_PATH, "w", encoding="utf-8") as f:
        f.write(drop_number)

    print(drop_number)

if __name__ == "__main__":
    main()
