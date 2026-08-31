import openpyxl
import re
from difflib import SequenceMatcher
from collections import Counter

# ─────────────────────────────────────────────────────────────────────
# NOTE ON GENERICIZATION
# Client, employer, and system names have been replaced throughout this
# file for public release: CLIENT_A / CLIENT_B are two national accounts,
# and "the portal" is a proprietary service-management platform.
#
# The document-title keyword lists below are therefore illustrative
# rather than runnable — in production they contain the real strings that
# appear in scanned document titles. The matching logic, thresholds, and
# control flow are unchanged.
# ─────────────────────────────────────────────────────────────────────

# == HOW TO RUN ================================================================
# 1. Close Service_Ticket_Report.xlsx in Excel.
# 2. Open Command Prompt.
# 3. Type:  py C:\Automation\compare_equipment.py
# 4. Press Enter.
#
# WHAT IT DOES
#   Reads:   SOW_Lookup        - what each location's contract expects
#            WO_Counts_Detail  - what was actually scanned (already summed/WO)
#   Writes:  Equipment_Compliance  (one NEW tab; nothing else is touched)
#
#   Per location + month + device type (ERB / IRT / IFL) it shows the raw
#   numbers AND a verdict, so the branch can see exactly where a tech is
#   over- or under-servicing.
#
# COLUMNS in Equipment_Compliance:
#   ADDRESS, CLIENT, MONTH, MATCH_TYPE,
#   <ERB|IRT|IFL>_DEVICES         - devices on site (from SOW)
#   <ERB|IRT|IFL>_FREQUENCY       - how often each should be serviced (SOW)
#   <ERB|IRT|IFL>_EXPECTED_SCANS  - devices x visits-per-month
#   <ERB|IRT|IFL>_TOTAL_SCANS     - actual sum across the month's WOs
#   <ERB|IRT|IFL>_EFFECTIVE_VISITS - total_scans / devices (avg per device)
#   <ERB|IRT|IFL>_VERDICT         - PASS / UNDER (n short) / OVER (~Nx, exp Mx)
#                                   / CANCELLED / EXTRA DEVICE / NO SOW / N/A
#   OVERALL_VERDICT, FLAGS
#
# VERDICT THRESHOLDS (per device type):
#   total_scans within  95-110% of expected   -> PASS
#   total_scans <  95%  of expected           -> UNDER (n short)
#   total_scans > 110%  of expected           -> OVER  (~Nx, expected Mx)
# =============================================================================

TICKET_PATH = r"C:\Automation\Service_Ticket_Report.xlsx"
OUTPUT_TAB  = "Equipment_Compliance"

UNDER_PCT    = 0.95   # below this % of expected -> UNDER
OVER_PCT     = 1.10   # above this % of expected -> OVER
FUZZY_CUTOFF = 0.90   # min similarity to accept a fuzzy address match

# -- address normalizer (in-memory only; does not change any address) --------
SUFFIX = {'ROAD':'RD','STREET':'ST','AVENUE':'AVE','DRIVE':'DR','BOULEVARD':'BLVD',
          'LANE':'LN','HIGHWAY':'HWY','CIRCLE':'CIR','COURT':'CT','PARKWAY':'PKWY',
          'PLACE':'PL','TERRACE':'TER','ROUTE':'RT'}

def norm(a):
    if a is None: return ''
    s = str(a).split(',')[0].upper().replace('.', '')
    s = re.sub(r'\s+', ' ', s).strip()
    toks = s.split()
    for i, t in enumerate(toks):
        if t[:1].isdigit():
            toks = toks[i:]
            break
    return ' '.join(SUFFIX.get(w, w) for w in toks)

def to_num(v):
    if v is None or v == '': return None
    try: return float(v)
    except (ValueError, TypeError): return None  # e.g. literal "CANCELLED"

# == LOAD =====================================================================
print("=== Equipment Compliance: scanned vs expected ===\n")
wb = openpyxl.load_workbook(TICKET_PATH)

# SOW_Lookup columns: ADDRESS, ERB_COUNT, ERB_FREQ, ERB_VISITS, ERB_EXPECTED,
#                     IRT_COUNT, IRT_FREQ, IRT_VISITS, IRT_EXPECTED,
#                     IFL_COUNT, IFL_FREQ, IFL_VISITS, IFL_EXPECTED, FLAGS
sow = {}
for r in wb["SOW_Lookup"].iter_rows(min_row=2, values_only=True):
    if not r[0]: continue
    flags = str(r[13] or '').upper()
    sow[norm(r[0])] = {
        'display':   r[0],
        'cancelled': 'CANCEL' in flags,
        'erb_dev': to_num(r[1]),  'erb_freq': r[2],  'erb_exp': to_num(r[4]),
        'irt_dev': to_num(r[5]),  'irt_freq': r[6],  'irt_exp': to_num(r[8]),
        'ifl_dev': to_num(r[9]),  'ifl_freq': r[10], 'ifl_exp': to_num(r[12]),
        'erb_vis': to_num(r[3]),  'irt_vis': to_num(r[7]),  'ifl_vis': to_num(r[11]),
    }
sow_keys = list(sow)

# WO_Counts_Detail columns: WORK ORDER, ADDRESS, RAW_ADDRESS, TECHNICIAN,
#                           BRANCH, COMPLETED, MONTH, CLIENT, ACTIVITY,
#                           ERB, IRT, IFL, REVIEW_FLAGS
groups = {}
for r in wb["WO_Counts_Detail"].iter_rows(min_row=2, values_only=True):
    if not r[1] or not r[6]: continue
    key = norm(r[1]); month = str(r[6]).strip()
    g = groups.setdefault((key, month), {'erb':0,'irt':0,'ifl':0,
                                         'client':r[7],'display':r[1]})
    g['erb'] += r[9]  or 0
    g['irt'] += r[10] or 0
    g['ifl'] += r[11] or 0

# == MATCH + VERDICT ==========================================================
def find_sow(key):
    if key in sow: return sow[key], 'exact'
    best, score = None, 0.0
    for sk in sow_keys:
        s = SequenceMatcher(None, key, sk).ratio()
        if s > score:
            best, score = sk, s
    if score >= FUZZY_CUTOFF: return sow[best], 'fuzzy'
    return None, 'none'

def device_verdict(scan, exp, devices, expected_visits, cancelled, sow_found):
    """Returns (verdict_text, is_clean) where is_clean=True means counts toward PASS overall."""
    if cancelled:                     return 'CANCELLED', False
    if not sow_found:                 return 'NO SOW', False
    if exp is None or devices is None:
        return ('EXTRA DEVICE' if scan > 0 else 'NO SOW'), False
    if exp == 0:
        return ('EXTRA DEVICE' if scan > 0 else 'N/A'), (scan == 0)
    if scan < exp * UNDER_PCT:
        return f'UNDER ({int(round(exp - scan))} short)', False
    if scan > exp * OVER_PCT:
        eff = scan / devices if devices else 0
        exp_v = int(round(expected_visits)) if expected_visits else '?'
        return f'OVER (~{int(round(eff))}x, expected {exp_v}x)', False
    return 'PASS', True

def overall(v_erb, v_irt, v_ifl, mtype):
    parts = [v_erb, v_irt, v_ifl]
    if mtype == 'none' or all(p == 'NO SOW' for p in parts): return 'NO SOW (check map)'
    if 'CANCELLED' in parts:                       return 'CANCELLED'
    if any(p.startswith('UNDER')    for p in parts): return 'UNDER'
    if any(p.startswith('OVER')     for p in parts): return 'OVER'
    if 'EXTRA DEVICE' in parts:                    return 'CHECK (extra device)'
    return 'PASS'

# == BUILD ROWS ==============================================================
out_rows = []
for (key, month), g in groups.items():
    rec, mtype = find_sow(key)
    found = rec is not None
    cancelled = bool(rec and rec['cancelled'])

    def block(prefix, scan):
        if not found:
            return None, None, None, None, None, ('CANCELLED' if cancelled else 'NO SOW')
        devices = rec[f'{prefix}_dev'] if not cancelled else None
        freq    = rec[f'{prefix}_freq'] if not cancelled else ''
        exp     = rec[f'{prefix}_exp'] if not cancelled else None
        visits  = rec[f'{prefix}_vis'] if not cancelled else None
        eff_v   = round(scan / devices, 1) if (devices and devices > 0) else ''
        v, _    = device_verdict(scan, exp, devices, visits, cancelled, True)
        return devices, freq, exp, eff_v, v, None

    erb_dev, erb_freq, erb_exp, erb_eff, erb_v, _ = block('erb', g['erb']) if found else (None,None,None,None,'NO SOW',None)
    irt_dev, irt_freq, irt_exp, irt_eff, irt_v, _ = block('irt', g['irt']) if found else (None,None,None,None,'NO SOW',None)
    ifl_dev, ifl_freq, ifl_exp, ifl_eff, ifl_v, _ = block('ifl', g['ifl']) if found else (None,None,None,None,'NO SOW',None)

    if cancelled:
        erb_v = irt_v = ifl_v = 'CANCELLED'

    flag = ''
    if mtype == 'fuzzy': flag = f'FUZZY MATCH to "{rec["display"]}" - verify'

    out_rows.append([
        g['display'], g['client'], month, mtype,
        erb_dev or '', erb_freq or '', erb_exp if erb_exp is not None else '',
        g['erb'], erb_eff, erb_v,
        irt_dev or '', irt_freq or '', irt_exp if irt_exp is not None else '',
        g['irt'], irt_eff, irt_v,
        ifl_dev or '', ifl_freq or '', ifl_exp if ifl_exp is not None else '',
        g['ifl'], ifl_eff, ifl_v,
        overall(erb_v, irt_v, ifl_v, mtype), flag,
    ])

# Locations in SOW with NO work orders this period (cancelled? verify)
wo_keys = {k for (k, _m) in groups}
for key in sow:
    if key not in wo_keys and not sow[key]['cancelled']:
        rec = sow[key]
        out_rows.append([
            rec['display'], '', '(no work orders)', 'none',
            rec['erb_dev'] or '', rec['erb_freq'] or '', rec['erb_exp'] if rec['erb_exp'] is not None else '',
            '', '', 'NO WORK ORDERS',
            rec['irt_dev'] or '', rec['irt_freq'] or '', rec['irt_exp'] if rec['irt_exp'] is not None else '',
            '', '', 'NO WORK ORDERS',
            rec['ifl_dev'] or '', rec['ifl_freq'] or '', rec['ifl_exp'] if rec['ifl_exp'] is not None else '',
            '', '', 'NO WORK ORDERS',
            'NO WORK ORDERS - verify (cancelled?)', '',
        ])

out_rows.sort(key=lambda x: (str(x[1]), str(x[0]), str(x[2])))

# == WRITE ====================================================================
if OUTPUT_TAB in wb.sheetnames:
    del wb[OUTPUT_TAB]
ws = wb.create_sheet(OUTPUT_TAB)

headers = [
    'ADDRESS', 'CLIENT', 'MONTH', 'MATCH_TYPE',
    'ERB_DEVICES', 'ERB_FREQUENCY', 'ERB_EXPECTED_SCANS',
    'ERB_TOTAL_SCANS', 'ERB_EFFECTIVE_VISITS', 'ERB_VERDICT',
    'IRT_DEVICES', 'IRT_FREQUENCY', 'IRT_EXPECTED_SCANS',
    'IRT_TOTAL_SCANS', 'IRT_EFFECTIVE_VISITS', 'IRT_VERDICT',
    'IFL_DEVICES', 'IFL_FREQUENCY', 'IFL_EXPECTED_SCANS',
    'IFL_TOTAL_SCANS', 'IFL_EFFECTIVE_VISITS', 'IFL_VERDICT',
    'OVERALL_VERDICT', 'FLAGS',
]
ws.append(headers)
for row in out_rows: ws.append(row)
wb.save(TICKET_PATH)

# == SUMMARY ==================================================================
tally = Counter(r[22] for r in out_rows)
print(f"Rows written: {len(out_rows)}")
print("Overall verdict breakdown:")
for k, v in tally.most_common():
    print(f"  {k:<38} {v}")

print("\nSANITY CHECK - worked example, April 2026:")
print("  ERB  32 devices, 2x/mo ->  64 expected; 109 scanned = ~3x  -> OVER")
print("  IRT  54 devices, 1x/wk -> 216 expected; 212 scanned = 98%  -> PASS")
print("  IFL   2 devices, 1x/wk ->   8 expected;  12 scanned = 150% -> OVER")
print("\n=== Done. Wrote '%s' tab. Nothing else changed. ===" % OUTPUT_TAB)
