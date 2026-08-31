import sys
import re
import json
import openpyxl
from datetime import datetime, date

# ─────────────────────────────────────────────────────────────────────
# NOTE ON GENERICIZATION
# Client, employer, and system names have been replaced throughout this
# file for public release: CLIENT_A / CLIENT_B are two national accounts,
# and "the portal" is a proprietary service-management platform.
#
# The document-title keyword lists below are therefore illustrative
# rather than runnable — in production they contain the real strings that
# appear in scanned document titles. The matching logic, thresholds, and
# control flow are unchanged.
# ─────────────────────────────────────────────────────────────────────

# ── WHAT THIS SCRIPT DOES ─────────────────────────────────────────────
# Auto-called by Power Automate after each location scrape.
#
#   Normal location  -> writes one row per scraped document
#   Multi-drop tab   -> writes ONE "MANUAL REVIEW - MULTI DROP" marker row
#
# HISTORY IS PRESERVED. Only rows for THIS address pulled TODAY are replaced.
# Previous days' pulls stay put so day-over-day the portal discrepancies are
# visible. evaluate_compliance.py already filters to the most recent pull
# per address, so extra history does not affect compliance verdicts.

EXCEL_PATH = r"C:\Automation\Compliance_Model.xlsx"
SHEET_NAME = "Raw_Document_Pull"
MULTIDROP_SHEETS = ['Client_A_MultiDrop', 'Client_B_MultiDrop']

# ── TEMP FILE READER (encoding-tolerant) ──────────────────────────────
# PAD writes these files and is inconsistent about encoding — sometimes
# UTF-8, sometimes UTF-8-with-BOM, sometimes UTF-16. Try each in turn.
def read_temp_text(path, default=None):
    for enc in ('utf-8-sig', 'utf-16', 'latin-1'):
        try:
            with open(path, 'r', encoding=enc) as f:
                return f.read().strip()
        except FileNotFoundError:
            return default
        except (UnicodeDecodeError, UnicodeError):
            continue
    return default

# ── ADDRESS NORMALIZER ────────────────────────────────────────────────
# Identical logic to norm_addr() in evaluate_compliance.py so both scripts
# agree on which addresses are multi-drop. Do not change one without the other.
_SUFFIX = {'ROAD':'RD','STREET':'ST','AVENUE':'AVE','DRIVE':'DR','BOULEVARD':'BLVD',
           'LANE':'LN','HIGHWAY':'HWY','CIRCLE':'CIR','COURT':'CT','PARKWAY':'PKWY',
           'PLACE':'PL','TERRACE':'TER','ROUTE':'RT'}

def norm_addr(addr):
    if not addr: return ''
    s = str(addr).split(',')[0].upper().replace('.', '')
    s = re.sub(r'\s+', ' ', s).strip()
    toks = s.split()
    for i, t in enumerate(toks):
        if t[:1].isdigit():
            toks = toks[i:]
            break
    return ' '.join(_SUFFIX.get(w, w) for w in toks)

# ── DATE NORMALIZER (for comparing pull_date cells) ───────────────────
def to_date(v):
    """Turn whatever is in the pull_date cell into a date, or None."""
    if v is None: return None
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date):     return v
    s = str(v).strip()
    if not s: return None
    for fmt in ('%m/%d/%Y', '%m/%d/%y', '%Y-%m-%d'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None

# ── READ ADDRESS ──────────────────────────────────────────────────────
address = read_temp_text(r'C:\Automation\temp_address.txt')
if not address:
    print('ERROR: temp_address.txt not found or empty.')
    sys.exit(1)

# ── CALCULATE EXP STATUS ──────────────────────────────────────────────
def get_exp_status(exp_date_str):
    if not exp_date_str or exp_date_str == "N/A":
        return "NO_EXP_REQUIRED"
    if exp_date_str == "01/01/1900":
        return "NO_EXP_REQUIRED"
    try:
        exp_date = datetime.strptime(exp_date_str, "%m/%d/%Y")
        today = datetime.today()
        days_until = (exp_date - today).days
        if days_until < 0:
            return "EXPIRED"
        elif days_until <= 30:
            return "EXPIRING_SOON"
        else:
            return "VALID"
    except ValueError:
        return "NO_EXP_REQUIRED"

# ── OPEN EXCEL ────────────────────────────────────────────────────────
try:
    wb = openpyxl.load_workbook(EXCEL_PATH)
except FileNotFoundError:
    print(f"ERROR: Excel file not found at {EXCEL_PATH}")
    sys.exit(1)
except PermissionError:
    print("ERROR: Excel file is open. Close it completely and re-run.")
    sys.exit(1)

if SHEET_NAME not in wb.sheetnames:
    print(f"ERROR: Sheet '{SHEET_NAME}' not found in workbook.")
    sys.exit(1)

# ── MULTI-DROP LOOKUP (from the workbook, NOT temp_dropnumber.txt) ────
# Source of truth = the Client_A_MultiDrop / Client_B_MultiDrop tabs.
# Column A = address, Column B = drop number. Drop 2 or 3 means the correct
# the portal result is NOT the first one, so the scrape can't be trusted.
# Drop 1 (or absent) = ordinary location, runs normally even if the portal
# happens to show more than one search result.
multidrop_addrs = set()
sheets_found = []
for sheet in MULTIDROP_SHEETS:
    if sheet in wb.sheetnames:
        sheets_found.append(sheet)
        for i, row in enumerate(wb[sheet].iter_rows(values_only=True)):
            if i == 0: continue          # header
            if not row or not row[0]: continue
            drop_num = str(row[1]).strip() if len(row) > 1 and row[1] is not None else '1'
            if drop_num in ('2', '3'):
                multidrop_addrs.add(norm_addr(row[0]))

if not sheets_found:
    print("WARNING: No MultiDrop tabs found — treating every location as normal.")

is_multidrop = norm_addr(address) in multidrop_addrs

ws = wb[SHEET_NAME]
pull_date = datetime.today().strftime("%m/%d/%Y")
today = date.today()

# ── REPLACE ONLY TODAY'S ROWS FOR THIS ADDRESS ────────────────────────
# History from previous pull dates is deliberately left alone — that record
# is the point of this tab. Re-running the same address on the same day
# replaces that day's rows instead of duplicating them.
rows_to_delete = []
for row_num in range(2, ws.max_row + 1):
    cell_addr = ws.cell(row=row_num, column=1).value   # A = Address
    cell_pull = ws.cell(row=row_num, column=6).value   # F = pull_date
    if not cell_addr:
        continue
    if str(cell_addr).strip().upper() != address.strip().upper():
        continue
    if to_date(cell_pull) == today:
        rows_to_delete.append(row_num)

for row_num in reversed(rows_to_delete):
    ws.delete_rows(row_num, 1)

if rows_to_delete:
    print(f"  Replaced {len(rows_to_delete)} row(s) from today's earlier pull. Prior days kept.")

# ── MULTI-DROP: one manual-review marker row, then stop ───────────────
if is_multidrop:
    ws.append([
        address,                        # A - Address
        "multidrop",                    # B - doc_area
        "N/A",                          # C - tech_name
        "MANUAL REVIEW - MULTI DROP",   # D - doc_title
        "N/A",                          # E - exp_date
        pull_date,                      # F - pull_date
        "MANUAL_REVIEW",                # G - exp_status
        "MANUAL_REVIEW",                # H - name_match
        "N/A"                           # I - location_specific
    ])
    wb.save(EXCEL_PATH)
    print(f"MULTI-DROP (found in MultiDrop tab): marker written for {address}")
    sys.exit(0)

# ── NORMAL LOCATION: one row per scraped document ─────────────────────
raw_json = read_temp_text(r"C:\Automation\temp_docs.json")
if raw_json is None:
    print("ERROR: temp_docs.json not found.")
    sys.exit(1)

try:
    documents = json.loads(raw_json)
except json.JSONDecodeError as e:
    print(f"ERROR: Could not parse DocumentData JSON: {e}")
    sys.exit(1)

rows_written = 0
for doc in documents:
    address_val       = address
    doc_area          = doc.get("table_type", "N/A")
    tech_name         = doc.get("employee_name", "N/A")
    doc_title         = doc.get("document", "N/A")
    exp_date          = doc.get("expiration", "N/A")
    location_specific = doc.get("location_specific", "N/A")
    inspection_date   = doc.get("inspection_date", "")

    # For qa_manual entries use inspection_date as a note in doc_title if present
    if doc_area == "qa_manual" and inspection_date:
        doc_title = doc_title  # keep as is — inspection_date stored separately
        tech_name = inspection_date  # store inspection date in tech_name field for qa_manual

    exp_status = get_exp_status(exp_date)

    ws.append([
        address_val,       # A - Address
        doc_area,          # B - doc_area
        tech_name,         # C - tech_name
        doc_title,         # D - doc_title
        exp_date,          # E - exp_date
        pull_date,         # F - pull_date
        exp_status,        # G - exp_status
        "PENDING",         # H - name_match (Python compliance layer will fill this)
        location_specific  # I - location_specific
    ])
    rows_written += 1

wb.save(EXCEL_PATH)
print(f"SUCCESS: {rows_written} rows written for {address} (pull date {pull_date})")
